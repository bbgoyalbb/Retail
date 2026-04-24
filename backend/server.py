from fastapi import FastAPI, APIRouter, HTTPException, Query, UploadFile, File, Depends, Header, Body, status, Request
from fastapi.responses import StreamingResponse, JSONResponse, FileResponse
from fastapi.staticfiles import StaticFiles
from fastapi.security import HTTPAuthorizationCredentials
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from starlette.middleware.gzip import GZipMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
import io
import json
from pathlib import Path
from pydantic import BaseModel, Field
from typing import List, Optional, Dict, Any
import uuid
from datetime import datetime, timezone, date
from bson import ObjectId
import re
from data_quality import (
    round_money as dq_round_money,
    determine_payment_status as dq_determine_payment_status,
    build_payment_mode_label as dq_build_payment_mode_label,
    generate_data_audit as dq_generate_data_audit,
    normalize_low_risk_data as dq_normalize_low_risk_data,
    repair_high_risk_data as dq_repair_high_risk_data,
)
import auth as auth_module

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

app = FastAPI()

api_router = APIRouter(prefix="/api")

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# Suppress WinError 10054 noise (browser forcibly closes SSL connections on Windows)
logging.getLogger("asyncio").setLevel(logging.CRITICAL)

# --- Auth dependency ---
async def get_current_user_dep(credentials: HTTPAuthorizationCredentials = Depends(auth_module.security)):
    return await auth_module.get_current_user(credentials, db)

# --- Auth models ---
class LoginRequest(BaseModel):
    username: str
    password: str

class UserCreateRequest(BaseModel):
    username: str
    password: str
    full_name: str
    role: str = "cashier"
    allowed_pages: List[str] = []

# --- Rate limiting ---
_login_attempts: dict = {}
_RATE_LIMIT_MAX = 5
_RATE_LIMIT_WINDOW = 900  # 15 minutes

def _check_rate_limit(ip: str):
    now = datetime.now(timezone.utc).timestamp()
    attempts = [t for t in _login_attempts.get(ip, []) if now - t < _RATE_LIMIT_WINDOW]
    _login_attempts[ip] = attempts
    if len(attempts) >= _RATE_LIMIT_MAX:
        raise HTTPException(status_code=429, detail="Too many login attempts. Try again in 15 minutes.")
    _login_attempts[ip].append(now)

def _clear_rate_limit(ip: str):
    _login_attempts.pop(ip, None)

# ==========================================
# MODELS
# ==========================================

ARTICLE_TYPES = ["Shirt", "Pant", "Gurkha Pant", "Kurta", "Pajama", "Blazer", "Safari Shirt", "Indo", "Sherwani", "Jacket", "W Coat"]

TAILORING_RATES = {
    "Shirt": (500, 400), "Kurta": (500, 400),
    "Pant": (700, 500), "Pajama": (700, 500),
    "Gurkha Pant": (900, 600),
    "Blazer": (3500, 2150),
    "Safari Shirt": (1000, 600),
    "Indo": (4200, 2750), "Sherwani": (4200, 2750),
    "Jacket": (1700, 1100),
    "W Coat": (600, 600),
}

ADDON_ITEMS = ["Bow", "Tie", "Cufflinks", "Stall", "Buttons", "Saffa", "Dye", "Malla", "Kalangi"]

PAYMENT_MODES = ["Cash", "PhonePe", "Google Pay [E]", "Google Pay [S]", "Bank Transfer"]

class BillLineItem(BaseModel):
    barcode: str
    qty: float
    price: float
    discount: float = 0

class CreateBillRequest(BaseModel):
    customer_name: str
    date: str
    payment_date: str
    items: List[BillLineItem]
    payment_modes: List[str] = ["Cash"]
    amount_paid: float = 0
    is_settled: bool = False
    needs_tailoring: bool = False

class TailoringOrderRequest(BaseModel):
    item_ids: List[str]
    order_no: str
    delivery_date: str
    assignments: List[dict]  # [{item_id, article_type, embroidery_status, split_data?}]

class AddOnRequest(BaseModel):
    item_id: str
    addons: List[dict]  # [{name, price}]

class StatusUpdateRequest(BaseModel):
    item_ids: List[str]
    new_status: str
    karigar: Optional[str] = None

class SettlementRequest(BaseModel):
    customer_name: str
    ref: str
    payment_date: str
    payment_modes: List[str]
    fresh_payment: float = 0
    use_advance: bool = False
    allot_fabric: float = 0
    allot_tailoring: float = 0
    allot_embroidery: float = 0
    allot_addon: float = 0
    allot_advance: float = 0

class TallyRequest(BaseModel):
    entry_ids: List[str]
    category: str
    action: str  # "tally" or "untally"
    date: Optional[str] = None  # scope tally to a specific pay-date row

class LabourPaymentRequest(BaseModel):
    item_ids: List[str]
    labour_type: str  # "tailoring" or "embroidery"
    payment_date: str
    payment_modes: List[str]
    payment_id: Optional[str] = None

class ItemUpdateRequest(BaseModel):
    barcode: Optional[str] = None
    price: Optional[float] = None
    qty: Optional[float] = None
    discount: Optional[float] = None
    fabric_amount: Optional[float] = None
    name: Optional[str] = None
    date: Optional[str] = None
    tailoring_status: Optional[str] = None
    article_type: Optional[str] = None
    order_no: Optional[str] = None
    delivery_date: Optional[str] = None
    tailoring_amount: Optional[float] = None
    embroidery_status: Optional[str] = None
    embroidery_amount: Optional[float] = None
    addon_desc: Optional[str] = None
    addon_amount: Optional[float] = None
    fabric_pay_mode: Optional[str] = None
    fabric_pending: Optional[float] = None
    fabric_received: Optional[float] = None
    tailoring_pay_mode: Optional[str] = None
    tailoring_pending: Optional[float] = None
    tailoring_received: Optional[float] = None
    embroidery_pay_mode: Optional[str] = None
    embroidery_pending: Optional[float] = None
    embroidery_received: Optional[float] = None
    karigar: Optional[str] = None

# ==========================================
# HELPERS
# ==========================================

def make_ref(seq: int, date_str: str) -> str:
    try:
        parts = date_str.split("-")
        if len(parts) == 3:
            d, m, y = parts
            return f"{seq:02d}/{d}{m}{y[2:]}"
    except Exception:
        pass
    return f"{seq:02d}/000000"

def serialize_doc(doc):
    if doc is None:
        return None
    doc["_id"] = str(doc["_id"])
    return doc

def serialize_docs(docs):
    return [serialize_doc(d) for d in docs]

def round_money(value: float) -> float:
    return round(float(value or 0), 2)

def determine_payment_status(pending_amount: float, received_amount: float) -> str:
    pending_amount = round_money(pending_amount)
    received_amount = round_money(received_amount)
    if received_amount > 0:
        return "Settled"
    if pending_amount > 0:
        return "Pending"
    return "N/A"

def build_payment_mode_label(payment_modes: List[str], pending_amount: float, received_amount: float) -> str:
    status = determine_payment_status(pending_amount, received_amount)
    if status == "Settled":
        return f"Settled - {', '.join(payment_modes) if payment_modes else 'Cash'}"
    if status == "Pending":
        return "Pending"
    return "N/A"

def analyze_payment_field(
    item: dict,
    amount_field: str,
    received_field: str,
    pending_field: str,
    mode_field: str,
    label: str,
) -> List[dict]:
    issues = []
    total = round_money(item.get(amount_field, 0))
    received = round_money(item.get(received_field, 0))
    pending = round_money(item.get(pending_field, 0))
    mode = item.get(mode_field, "N/A") or "N/A"
    expected_status = determine_payment_status(pending, received)

    # negative_pending and received>total are intentional when an over-payment was made.
    # Flag as informational 'overpaid' rather than an error so audit still surfaces them.
    if pending < 0 or (total >= 0 and received - total > 0.01):
        issues.append({
            "type": "overpaid",
            "category": label,
            "message": f"{label} over-payment: received ₹{received} against total ₹{total} (credit ₹{round_money(received - total)})",
            "total": total,
            "received": received,
            "pending": pending,
            "mode": mode,
        })

    if received < 0:
        issues.append({
            "type": "negative_received",
            "category": label,
            "message": f"{label} received is negative",
            "total": total,
            "received": received,
            "pending": pending,
            "mode": mode,
        })

    if total >= 0 and pending - total > 0.01:
        issues.append({
            "type": "pending_exceeds_total",
            "category": label,
            "message": f"{label} pending exceeds total amount",
            "total": total,
            "received": received,
            "pending": pending,
            "mode": mode,
        })

    # Only flag amount_mismatch when pending >= 0 (negative pending is an intentional over-payment state)
    if total > 0 and pending >= 0 and abs(round_money(received + pending) - total) > 1:
        issues.append({
            "type": "amount_mismatch",
            "category": label,
            "message": f"{label} total does not equal received plus pending",
            "total": total,
            "received": received,
            "pending": pending,
            "mode": mode,
        })

    if expected_status == "Pending" and mode != "Pending":
        issues.append({
            "type": "mode_status_mismatch",
            "category": label,
            "message": f"{label} is pending but mode is {mode}",
            "total": total,
            "received": received,
            "pending": pending,
            "mode": mode,
        })

    if expected_status == "Settled" and not str(mode).startswith("Settled"):
        issues.append({
            "type": "mode_status_mismatch",
            "category": label,
            "message": f"{label} is settled but mode is {mode}",
            "total": total,
            "received": received,
            "pending": pending,
            "mode": mode,
        })

    return issues

def normalize_payment_field(
    item: dict,
    amount_field: str,
    received_field: str,
    pending_field: str,
    mode_field: str,
) -> dict:
    total = round_money(item.get(amount_field, 0))
    received = round_money(item.get(received_field, 0))
    pending = round_money(item.get(pending_field, 0))
    original_mode = item.get(mode_field, "N/A") or "N/A"

    if pending < 0 and abs(pending) <= 1:
        pending = 0.0

    if received < 0 and abs(received) <= 1:
        received = 0.0

    if total >= 0 and received > total and abs(received - total) <= 1:
        received = total

    if total >= 0 and pending > total and abs(pending - total) <= 1:
        pending = total

    if total > 0:
        mismatch = round_money((received + pending) - total)
        if abs(mismatch) <= 1:
            if pending > 0:
                pending = round_money(max(0, total - received))
            else:
                received = round_money(max(0, total - pending))

    status = determine_payment_status(pending, received)
    mode = original_mode
    if original_mode != "N/A" or total > 0 or received > 0 or pending > 0:
        mode = status if status != "N/A" else "N/A"
        if status == "Settled":
            mode_suffix = ""
            if " - " in str(original_mode):
                mode_suffix = original_mode.split(" - ", 1)[1].strip()
                if mode_suffix.startswith("Partially Settled - "):
                    mode_suffix = mode_suffix[len("Partially Settled - "):].strip()
            if mode_suffix:
                mode = f"Settled - {mode_suffix}"

    return {
        received_field: received,
        pending_field: pending,
        mode_field: mode,
    }

async def normalize_low_risk_data(limit: int = 100) -> dict:
    items = await db.items.find({}, {"_id": 0}).to_list(10000)
    advances = await db.advances.find({}, {"_id": 0}).to_list(5000)

    changes = []
    items_updated = 0
    advances_updated = 0

    for item in items:
        updates = {}
        checks = [
            ("fabric_amount", "fabric_received", "fabric_pending", "fabric_pay_mode"),
            ("tailoring_amount", "tailoring_received", "tailoring_pending", "tailoring_pay_mode"),
            ("embroidery_amount", "embroidery_received", "embroidery_pending", "embroidery_pay_mode"),
            ("addon_amount", "addon_received", "addon_pending", "addon_pay_mode"),
        ]

        for check in checks:
            normalized = normalize_payment_field(item, *check)
            for field, value in normalized.items():
                current_value = item.get(field)
                if isinstance(value, (int, float)):
                    changed = round_money(current_value) != round_money(value)
                else:
                    changed = current_value != value
                if changed:
                    updates[field] = value

        if updates:
            await db.items.update_one({"id": item["id"]}, {"$set": updates})
            items_updated += 1
            if len(changes) < limit:
                changes.append({
                    "kind": "item",
                    "item_id": item.get("id"),
                    "ref": item.get("ref"),
                    "name": item.get("name"),
                    "barcode": item.get("barcode"),
                    "updates": updates,
                })

    advance_totals = {}
    for adv in advances:
        ref = adv.get("ref", "")
        advance_totals[ref] = round_money(advance_totals.get(ref, 0) + round_money(adv.get("amount", 0)))

    for ref, total in advance_totals.items():
        if total < -0.01:
            negative_entries = [a for a in advances if a.get("ref") == ref and round_money(a.get("amount", 0)) < 0 and a.get("mode") != "Adjusted"]
            for adv in negative_entries:
                await db.advances.update_one({"id": adv["id"]}, {"$set": {"mode": "Adjusted"}})
                advances_updated += 1
                if len(changes) < limit:
                    changes.append({
                        "kind": "advance",
                        "advance_id": adv.get("id"),
                        "ref": ref,
                        "name": adv.get("name"),
                        "updates": {"mode": "Adjusted"},
                    })

    return {
        "items_updated": items_updated,
        "advances_updated": advances_updated,
        "changes": changes,
        "audit_after": await generate_data_audit(limit),
    }

async def repair_high_risk_data(limit: int = 100) -> dict:
    items = await db.items.find({}, {"_id": 0}).to_list(10000)
    item_updates = 0
    advances_created = 0
    changes = []

    for item in items:
        updates = {}
        carry_forwards = []
        checks = [
            ("fabric", "fabric_amount", "fabric_received", "fabric_pending", "fabric_pay_mode", "fabric_pay_date"),
            ("tailoring", "tailoring_amount", "tailoring_received", "tailoring_pending", "tailoring_pay_mode", "tailoring_pay_date"),
            ("embroidery", "embroidery_amount", "embroidery_received", "embroidery_pending", "embroidery_pay_mode", "embroidery_pay_date"),
            ("addon", "addon_amount", "addon_received", "addon_pending", "addon_pay_mode", "addon_pay_date"),
        ]

        for label, amount_field, received_field, pending_field, mode_field, date_field in checks:
            total = round_money(item.get(amount_field, 0))
            received = round_money(item.get(received_field, 0))
            pending = round_money(item.get(pending_field, 0))
            original_mode = item.get(mode_field, "N/A") or "N/A"

            if total <= 0 and received <= 0 and pending <= 0:
                continue

            excess = 0.0
            corrected_received = received
            corrected_pending = pending

            # Skip over-paid items — negative pending / received>total is intentional.
            # The repair tool must not undo deliberate over-payments.
            if received > total + 0.01 or corrected_pending < -0.01:
                continue

            if corrected_pending >= 0 and corrected_received <= total + 0.01:
                corrected_pending = round_money(max(0, total - corrected_received))

            corrected_mode = original_mode
            corrected_status = determine_payment_status(corrected_pending, corrected_received)
            if corrected_status == "Pending":
                corrected_mode = "Pending"
            elif corrected_status == "Settled":
                suffix = ""
                if " - " in str(original_mode):
                    suffix = original_mode.split(" - ", 1)[1].strip()
                    if suffix.startswith("Partially Settled - "):
                        suffix = suffix[len("Partially Settled - "):].strip()
                corrected_mode = f"Settled - {suffix}" if suffix else "Settled"
            elif total <= 0:
                corrected_mode = "N/A"

            field_updates = {
                received_field: corrected_received,
                pending_field: corrected_pending,
                mode_field: corrected_mode,
            }

            changed_fields = {}
            for field, value in field_updates.items():
                current_value = item.get(field)
                if isinstance(value, (int, float)):
                    changed = round_money(current_value) != round_money(value)
                else:
                    changed = current_value != value
                if changed:
                    changed_fields[field] = value
                    updates[field] = value

            if excess > 0.01:
                carry_forwards.append({
                    "category": label,
                    "amount": excess,
                    "date": item.get(date_field) if item.get(date_field) and item.get(date_field) != "N/A" else item.get("date"),
                })
                changed_fields["carry_forward"] = excess

            if changed_fields and len(changes) < limit:
                changes.append({
                    "kind": "item_repair",
                    "item_id": item.get("id"),
                    "ref": item.get("ref"),
                    "name": item.get("name"),
                    "barcode": item.get("barcode"),
                    "category": label,
                    "updates": changed_fields,
                })

        if updates:
            await db.items.update_one({"id": item["id"]}, {"$set": updates})
            item_updates += 1

        for carry in carry_forwards:
            adv = {
                "id": str(uuid.uuid4()),
                "date": carry["date"] or item.get("date"),
                "name": item.get("name", ""),
                "ref": item.get("ref", ""),
                "amount": carry["amount"],
                "mode": f"Auto Carry Forward - {carry['category'].title()}",
                "tally": False,
                "created_at": datetime.now(timezone.utc).isoformat(),
            }
            await db.advances.insert_one(adv)
            advances_created += 1
            if len(changes) < limit:
                changes.append({
                    "kind": "advance_created",
                    "ref": adv["ref"],
                    "name": adv["name"],
                    "amount": adv["amount"],
                    "mode": adv["mode"],
                })

    return {
        "items_updated": item_updates,
        "advances_created": advances_created,
        "changes": changes,
        "audit_after": await generate_data_audit(limit),
    }

async def generate_data_audit(limit: int = 100) -> dict:
    items = await db.items.find({}, {"_id": 0}).to_list(10000)
    advances = await db.advances.find({}, {"_id": 0}).to_list(5000)

    issue_counts = {}
    issues = []

    def push_issue(issue: dict):
        issue_counts[issue["type"]] = issue_counts.get(issue["type"], 0) + 1
        if len(issues) < limit:
            issues.append(issue)

    for item in items:
        base_info = {
            "item_id": item.get("id"),
            "ref": item.get("ref"),
            "name": item.get("name"),
            "barcode": item.get("barcode"),
            "date": item.get("date"),
        }

        checks = [
            ("fabric_amount", "fabric_received", "fabric_pending", "fabric_pay_mode", "fabric"),
            ("tailoring_amount", "tailoring_received", "tailoring_pending", "tailoring_pay_mode", "tailoring"),
            ("embroidery_amount", "embroidery_received", "embroidery_pending", "embroidery_pay_mode", "embroidery"),
            ("addon_amount", "addon_received", "addon_pending", "addon_pay_mode", "addon"),
        ]

        for check in checks:
            for issue in analyze_payment_field(item, *check):
                push_issue({**base_info, **issue})

        emb_labour = round_money(item.get("emb_labour_amount", 0))
        if emb_labour > 0 and item.get("embroidery_status") not in ["Finished", "In Progress"]:
            push_issue({
                **base_info,
                "type": "embroidery_labour_status_mismatch",
                "category": "embroidery_labour",
                "message": "Embroidery labour exists while embroidery status is not in progress/finished",
                "emb_labour_amount": emb_labour,
                "embroidery_status": item.get("embroidery_status"),
            })

    advance_total_by_ref = {}
    for adv in advances:
        ref = adv.get("ref", "")
        amount = round_money(adv.get("amount", 0))
        advance_total_by_ref[ref] = round_money(advance_total_by_ref.get(ref, 0) + amount)
        if amount < 0 and adv.get("mode") != "Adjusted":
            push_issue({
                "ref": ref,
                "name": adv.get("name"),
                "type": "negative_advance_non_adjustment",
                "category": "advance",
                "message": "Negative advance entry is not marked as Adjusted",
                "amount": amount,
                "mode": adv.get("mode"),
                "date": adv.get("date"),
            })

    for ref, total in advance_total_by_ref.items():
        if total < -0.01:
            push_issue({
                "ref": ref,
                "type": "negative_advance_balance",
                "category": "advance",
                "message": "Advance balance is negative for this reference",
                "amount": total,
            })

    return {
        "scanned": {
            "items": len(items),
            "advances": len(advances),
        },
        "total_issues": sum(issue_counts.values()),
        "issue_counts": dict(sorted(issue_counts.items(), key=lambda kv: (-kv[1], kv[0]))),
        "issues": issues,
    }

# ==========================================
# SEED DATA
# ==========================================

@api_router.post("/seed")
async def seed_data(current_user: dict = Depends(get_current_user_dep)):
    count = await db.items.count_documents({})
    if count > 0:
        return {"message": "Data already seeded", "items_count": count}

    try:
        import openpyxl
        wb_path = "/tmp/retail_book.xlsm"
        if not os.path.exists(wb_path):
            return {"message": "Excel file not found at /tmp/retail_book.xlsm"}

        wb = openpyxl.load_workbook(wb_path, data_only=True)

        ws = wb['Item Details']
        items = []
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
            if not row[0]:
                continue
            dt = row[0]
            date_str = dt.strftime("%Y-%m-%d") if hasattr(dt, 'strftime') else str(dt)[:10]

            fab_pay_date = ""
            if row[18] and hasattr(row[18], 'strftime'):
                fab_pay_date = row[18].strftime("%Y-%m-%d")

            delivery_date = ""
            if row[11] and hasattr(row[11], 'strftime'):
                delivery_date = row[11].strftime("%Y-%m-%d")

            tail_pay_date = ""
            if row[25] and hasattr(row[25], 'strftime'):
                tail_pay_date = row[25].strftime("%Y-%m-%d")

            emb_pay_date = ""
            if row[29] and hasattr(row[29], 'strftime'):
                emb_pay_date = row[29].strftime("%Y-%m-%d")

            addon_pay_date = ""
            if row[33] and hasattr(row[33], 'strftime'):
                addon_pay_date = row[33].strftime("%Y-%m-%d")

            labour_pay_date = ""
            if row[23] and hasattr(row[23], 'strftime'):
                labour_pay_date = row[23].strftime("%Y-%m-%d")

            def safe_float(v):
                if v is None or str(v).strip() in ("N/A", "", "None"):
                    return 0
                try:
                    return float(v)
                except (ValueError, TypeError):
                    return 0

            def safe_str(v):
                if v is None:
                    return "N/A"
                return str(v).strip()

            item = {
                "id": str(uuid.uuid4()),
                "date": date_str,
                "name": safe_str(row[1]),
                "ref": safe_str(row[2]),
                "barcode": safe_str(row[3]),
                "price": safe_float(row[4]),
                "qty": safe_float(row[5]),
                "discount": safe_float(row[6]),
                "fabric_amount": safe_float(row[7]),
                "tailoring_status": safe_str(row[8]),
                "article_type": safe_str(row[9]),
                "order_no": safe_str(row[10]),
                "delivery_date": delivery_date if delivery_date else "N/A",
                "tailoring_amount": safe_float(row[12]),
                "embroidery_status": safe_str(row[13]),
                "embroidery_amount": safe_float(row[14]),
                "addon_desc": safe_str(row[15]),
                "addon_amount": safe_float(row[16]),
                "fabric_pay_mode": safe_str(row[17]),
                "fabric_pay_date": fab_pay_date if fab_pay_date else "N/A",
                "fabric_pending": safe_float(row[19]),
                "fabric_received": safe_float(row[20]),
                "labour_amount": safe_float(row[21]),
                "labour_paid": safe_str(row[22]),
                "labour_pay_date": labour_pay_date if labour_pay_date else "N/A",
                "tailoring_pay_mode": safe_str(row[24]),
                "tailoring_pay_date": tail_pay_date if tail_pay_date else "N/A",
                "tailoring_received": safe_float(row[26]),
                "tailoring_pending": safe_float(row[27]),
                "embroidery_pay_mode": safe_str(row[28]),
                "embroidery_pay_date": emb_pay_date if emb_pay_date else "N/A",
                "embroidery_received": safe_float(row[30]),
                "embroidery_pending": safe_float(row[31]),
                "addon_pay_mode": safe_str(row[32]),
                "addon_pay_date": addon_pay_date if addon_pay_date else "N/A",
                "addon_received": safe_float(row[34]),
                "addon_pending": safe_float(row[35]),
                "karigar": safe_str(row[36]) if len(row) > 36 else "N/A",
                "tally_fabric": False,
                "tally_tailoring": False,
                "tally_embroidery": False,
                "tally_addon": False,
                "created_at": datetime.now(timezone.utc).isoformat(),
            }
            items.append(item)

        if items:
            await db.items.insert_many(items)

        ws2 = wb['Advances']
        advances = []
        for row in ws2.iter_rows(min_row=2, max_row=ws2.max_row, values_only=True):
            if not row[0]:
                continue
            dt = row[0]
            date_str = dt.strftime("%Y-%m-%d") if hasattr(dt, 'strftime') else str(dt)[:10]
            adv = {
                "id": str(uuid.uuid4()),
                "date": date_str,
                "name": str(row[1]).strip() if row[1] else "",
                "ref": str(row[2]).strip() if row[2] else "",
                "amount": float(row[3]) if row[3] else 0,
                "mode": str(row[4]).strip() if row[4] else "",
                "tally": False,
                "created_at": datetime.now(timezone.utc).isoformat(),
            }
            advances.append(adv)

        if advances:
            await db.advances.insert_many(advances)

        return {"message": "Data seeded successfully", "items_count": len(items), "advances_count": len(advances)}
    except Exception as e:
        logger.error(f"Seed error: {e}")
        raise HTTPException(status_code=500, detail=str(e))

# ==========================================
# DASHBOARD
# ==========================================

@api_router.get("/dashboard")
async def get_dashboard():
    total_items = await db.items.count_documents({})
    total_advances = await db.advances.count_documents({})

    _not_settled = {"$not": {"$regex": "^Settled"}}
    pipeline_fabric_pending = [
        {"$match": {"fabric_amount": {"$gt": 0}, "fabric_pay_mode": _not_settled}},
        {"$group": {"_id": None, "total": {"$sum": "$fabric_pending"}}}
    ]
    fab_pending = await db.items.aggregate(pipeline_fabric_pending).to_list(1)

    pipeline_tail_pending = [
        {"$match": {"tailoring_amount": {"$gt": 0}, "tailoring_pay_mode": _not_settled}},
        {"$group": {"_id": None, "total": {"$sum": "$tailoring_pending"}}}
    ]
    tail_pending = await db.items.aggregate(pipeline_tail_pending).to_list(1)

    pipeline_emb_pending = [
        {"$match": {"embroidery_amount": {"$gt": 0}, "embroidery_pay_mode": _not_settled}},
        {"$group": {"_id": None, "total": {"$sum": "$embroidery_pending"}}}
    ]
    emb_pending = await db.items.aggregate(pipeline_emb_pending).to_list(1)

    tailoring_pending_count = await db.items.count_documents({"tailoring_status": "Pending"})
    tailoring_stitched_count = await db.items.count_documents({"tailoring_status": "Stitched"})
    emb_required = await db.items.count_documents({"embroidery_status": "Required"})
    emb_inprogress = await db.items.count_documents({"embroidery_status": "In Progress"})

    unique_customers = await db.items.distinct("name")

    pipeline_revenue = [
        {"$group": {"_id": None, "total": {"$sum": "$fabric_received"}}}
    ]
    revenue = await db.items.aggregate(pipeline_revenue).to_list(1)

    pipeline_adv_total = [
        {"$group": {"_id": None, "total": {"$sum": "$amount"}}}
    ]
    adv_total = await db.advances.aggregate(pipeline_adv_total).to_list(1)

    recent_items = await db.items.find({}, {"_id": 0}).sort("date", -1).limit(10).to_list(10)

    return {
        "total_items": total_items,
        "total_advances": total_advances,
        "fabric_pending_amount": fab_pending[0]["total"] if fab_pending else 0,
        "tailoring_pending_amount": tail_pending[0]["total"] if tail_pending else 0,
        "embroidery_pending_amount": emb_pending[0]["total"] if emb_pending else 0,
        "tailoring_pending_count": tailoring_pending_count,
        "tailoring_stitched_count": tailoring_stitched_count,
        "embroidery_required_count": emb_required,
        "embroidery_inprogress_count": emb_inprogress,
        "unique_customers": len(unique_customers),
        "total_revenue": revenue[0]["total"] if revenue else 0,
        "total_advances_amount": adv_total[0]["total"] if adv_total else 0,
        "recent_items": recent_items,
    }

# ==========================================
# CUSTOMERS
# ==========================================

@api_router.get("/customers")
async def get_customers():
    customers = await db.items.distinct("name")
    return sorted([c for c in customers if c and c != "N/A"])

# ==========================================
# ITEMS CRUD
# ==========================================

@api_router.get("/items")
async def get_items(
    name: Optional[str] = None,
    ref: Optional[str] = None,
    date: Optional[str] = None,
    tailoring_status: Optional[str] = None,
    embroidery_status: Optional[str] = None,
    order_no: Optional[str] = None,
    limit: int = 500,
    skip: int = 0,
):
    query = {}
    if name:
        query["name"] = name
    if ref:
        query["ref"] = ref
    if date:
        query["date"] = date
    if tailoring_status:
        query["tailoring_status"] = tailoring_status
    if embroidery_status:
        query["embroidery_status"] = embroidery_status
    if order_no:
        query["order_no"] = order_no

    items = await db.items.find(query, {"_id": 0}).sort("date", -1).skip(skip).limit(limit).to_list(limit)
    total = await db.items.count_documents(query)
    return {"items": items, "total": total}

@api_router.get("/items/{item_id}")
async def get_item(item_id: str):
    item = await db.items.find_one({"id": item_id}, {"_id": 0})
    if not item:
        raise HTTPException(status_code=404, detail="Item not found")
    return item

@api_router.get("/refs")
async def get_refs(name: Optional[str] = None):
    query = {}
    if name:
        query["name"] = name
    pipeline = [
        {"$match": query},
        {"$group": {"_id": "$ref"}},
        {"$sort": {"_id": -1}}
    ]
    refs = await db.items.aggregate(pipeline).to_list(500)
    return [r["_id"] for r in refs if r["_id"] and r["_id"] != "N/A"]

# ==========================================
# NEW BILL
# ==========================================

@api_router.post("/bills")
async def create_bill(req: CreateBillRequest):
    if not req.items:
        raise HTTPException(status_code=400, detail="At least one item is required")

    existing = await db.items.find({"date": req.date}).distinct("ref")
    max_seq = 0
    try:
        parts = req.date.split("-")
        date_suffix = f"{parts[2]}{parts[1]}{parts[0][2:]}"
    except Exception:
        date_suffix = "000000"

    for r in existing:
        try:
            seq = int(r.split("/")[0])
            if seq > max_seq:
                max_seq = seq
        except Exception:
            pass

    ref = f"{max_seq + 1:02d}/{date_suffix}"
    modes_str = ", ".join(req.payment_modes) if req.payment_modes else "Cash"
    tailoring_status = "Awaiting Order" if req.needs_tailoring else "N/A"

    grand_total = 0
    for item in req.items:
        discounted_price = round(item.price - (item.price * item.discount / 100), 0)
        grand_total += round(discounted_price * item.qty, 0)

    if req.is_settled and req.amount_paid < grand_total:
        raise HTTPException(
            status_code=400,
            detail=f"Settled bills require full payment. Expected {grand_total:.0f}, received {req.amount_paid:.0f}."
        )

    items_to_insert = []
    running_paid = 0
    running_discount = 0

    for idx, item in enumerate(req.items):
        discounted_price = round(item.price - (item.price * item.discount / 100), 0)
        item_total = round(discounted_price * item.qty, 0)

        if req.is_settled:
            total_diff = grand_total - req.amount_paid
            if idx == len(req.items) - 1:
                item_discount = total_diff - running_discount
                item_paid = req.amount_paid - running_paid
            else:
                item_discount = round(item_total * (total_diff / grand_total), 0) if grand_total > 0 else 0
                item_paid = round(item_total * (req.amount_paid / grand_total), 0) if grand_total > 0 else 0
                running_discount += item_discount
                running_paid += item_paid

            doc = {
                "id": str(uuid.uuid4()),
                "date": req.date,
                "name": req.customer_name,
                "ref": ref,
                "barcode": item.barcode,
                "price": item.price,
                "qty": item.qty,
                "discount": item.discount,
                "fabric_amount": item_total,
                "tailoring_status": tailoring_status,
                "article_type": "N/A",
                "order_no": "N/A",
                "delivery_date": "N/A",
                "tailoring_amount": 0,
                "embroidery_status": "N/A",
                "embroidery_amount": 0,
                "addon_desc": "N/A",
                "addon_amount": 0,
                "fabric_pay_mode": f"Settled - {modes_str}",
                "fabric_pay_date": req.payment_date,
                "fabric_pending": item_discount,
                "fabric_received": item_paid,
                "labour_amount": 0,
                "labour_paid": "N/A",
                "labour_pay_date": "N/A",
                "tailoring_pay_mode": "N/A",
                "tailoring_pay_date": "N/A",
                "tailoring_received": 0,
                "tailoring_pending": 0,
                "embroidery_pay_mode": "N/A",
                "embroidery_pay_date": "N/A",
                "embroidery_received": 0,
                "embroidery_pending": 0,
                "addon_pay_mode": "N/A",
                "addon_pay_date": "N/A",
                "addon_received": 0,
                "addon_pending": 0,
                "karigar": "N/A",
                "tally_fabric": False,
                "tally_tailoring": False,
                "tally_embroidery": False,
                "tally_addon": False,
                "created_at": datetime.now(timezone.utc).isoformat(),
            }
        else:
            doc = {
                "id": str(uuid.uuid4()),
                "date": req.date,
                "name": req.customer_name,
                "ref": ref,
                "barcode": item.barcode,
                "price": item.price,
                "qty": item.qty,
                "discount": item.discount,
                "fabric_amount": item_total,
                "tailoring_status": tailoring_status,
                "article_type": "N/A",
                "order_no": "N/A",
                "delivery_date": "N/A",
                "tailoring_amount": 0,
                "embroidery_status": "N/A",
                "embroidery_amount": 0,
                "addon_desc": "N/A",
                "addon_amount": 0,
                "fabric_pay_mode": "Pending",
                "fabric_pay_date": "N/A",
                "fabric_pending": item_total,
                "fabric_received": 0,
                "labour_amount": 0,
                "labour_paid": "N/A",
                "labour_pay_date": "N/A",
                "tailoring_pay_mode": "N/A",
                "tailoring_pay_date": "N/A",
                "tailoring_received": 0,
                "tailoring_pending": 0,
                "embroidery_pay_mode": "N/A",
                "embroidery_pay_date": "N/A",
                "embroidery_received": 0,
                "embroidery_pending": 0,
                "addon_pay_mode": "N/A",
                "addon_pay_date": "N/A",
                "addon_received": 0,
                "addon_pending": 0,
                "karigar": "N/A",
                "tally_fabric": False,
                "tally_tailoring": False,
                "tally_embroidery": False,
                "tally_addon": False,
                "created_at": datetime.now(timezone.utc).isoformat(),
            }

        items_to_insert.append(doc)

    if items_to_insert:
        await db.items.insert_many(items_to_insert)

    if not req.is_settled and req.amount_paid > 0:
        adv = {
            "id": str(uuid.uuid4()),
            "date": req.payment_date,
            "name": req.customer_name,
            "ref": ref,
            "amount": req.amount_paid,
            "mode": modes_str,
            "tally": False,
            "created_at": datetime.now(timezone.utc).isoformat(),
        }
        await db.advances.insert_one(adv)

    return {"message": "Bill created", "ref": ref, "items_count": len(items_to_insert), "grand_total": grand_total}

# ==========================================
# TAILORING ORDERS
# ==========================================

@api_router.get("/tailoring/awaiting")
async def get_awaiting_orders():
    pipeline = [
        {"$match": {"tailoring_status": "Awaiting Order"}},
        {"$group": {
            "_id": {"name": "$name", "ref": "$ref"},
            "items": {"$push": {
                "id": "$id", "barcode": "$barcode", "price": "$price",
                "qty": "$qty", "article_type": "$article_type",
                "embroidery_status": "$embroidery_status"
            }},
            "date": {"$first": "$date"},
            "count": {"$sum": 1}
        }},
        {"$sort": {"date": -1}}
    ]
    result = await db.items.aggregate(pipeline).to_list(200)
    return [{"name": r["_id"]["name"], "ref": r["_id"]["ref"], "date": r["date"],
             "items": r["items"], "count": r["count"]} for r in result]

@api_router.post("/tailoring/assign")
async def assign_tailoring(req: TailoringOrderRequest):
    updated = 0
    for assignment in req.assignments:
        item_id = assignment.get("item_id")
        article_type = assignment.get("article_type", "Shirt")
        emb_status = assignment.get("embroidery_status", "Not Required")

        rates = TAILORING_RATES.get(article_type, (0, 0))
        tail_amt, labour_amt = rates

        update = {
            "$set": {
                "tailoring_status": "Pending",
                "article_type": article_type,
                "order_no": req.order_no,
                "delivery_date": req.delivery_date,
                "tailoring_amount": tail_amt,
                "tailoring_received": 0,
                "tailoring_pending": tail_amt,
                "tailoring_pay_mode": "Pending",
                "labour_amount": labour_amt,
                "embroidery_status": emb_status,
            }
        }
        if emb_status == "Required":
            update["$set"]["embroidery_pay_mode"] = "Pending"

        result = await db.items.update_one({"id": item_id}, update)
        if result.modified_count > 0:
            updated += 1

    return {"message": f"{updated} items assigned to order {req.order_no}"}

class SplitItem(BaseModel):
    article_type: str
    qty: float
    embroidery_status: str = "Not Required"

class SplitTailoringRequest(BaseModel):
    item_id: str
    order_no: str
    delivery_date: str
    splits: List[SplitItem]

@api_router.post("/tailoring/split")
async def split_and_assign(req: SplitTailoringRequest):
    item = await db.items.find_one({"id": req.item_id}, {"_id": 0})
    if not item:
        raise HTTPException(status_code=404, detail="Item not found")

    original_qty = item.get("qty", 0)
    original_price = item.get("price", 0)
    original_discount = item.get("discount", 0)

    created = 0
    for idx, split in enumerate(req.splits):
        rates = TAILORING_RATES.get(split.article_type, (0, 0))
        tail_amt, labour_amt = rates

        discounted_price = round(original_price - (original_price * original_discount / 100), 0)
        split_fabric_amt = round(discounted_price * split.qty, 0)

        if idx == 0:
            # Update original item with first split
            update = {
                "qty": split.qty,
                "fabric_amount": split_fabric_amt,
                "fabric_pending": split_fabric_amt if item.get("fabric_pay_mode") == "Pending" else item.get("fabric_pending", 0),
                "article_type": split.article_type,
                "tailoring_status": "Pending",
                "order_no": req.order_no,
                "delivery_date": req.delivery_date,
                "tailoring_amount": tail_amt,
                "tailoring_received": 0,
                "tailoring_pending": tail_amt,
                "tailoring_pay_mode": "Pending",
                "labour_amount": labour_amt,
                "embroidery_status": split.embroidery_status,
            }
            if split.embroidery_status == "Required":
                update["embroidery_pay_mode"] = "Pending"
            await db.items.update_one({"id": req.item_id}, {"$set": update})
        else:
            # Create new items for subsequent splits
            new_item = {**item}
            new_item.pop("_id", None)
            new_item["id"] = str(uuid.uuid4())
            new_item["qty"] = split.qty
            new_item["fabric_amount"] = split_fabric_amt
            new_item["fabric_pending"] = split_fabric_amt if item.get("fabric_pay_mode") == "Pending" else 0
            new_item["fabric_received"] = 0  # received stays on original item; new splits start at 0
            new_item["article_type"] = split.article_type
            new_item["tailoring_status"] = "Pending"
            new_item["order_no"] = req.order_no
            new_item["delivery_date"] = req.delivery_date
            new_item["tailoring_amount"] = tail_amt
            new_item["labour_amount"] = labour_amt
            new_item["tailoring_pending"] = tail_amt
            new_item["tailoring_pay_mode"] = "Pending"
            new_item["embroidery_status"] = split.embroidery_status
            if split.embroidery_status == "Required":
                new_item["embroidery_pay_mode"] = "Pending"
            new_item["created_at"] = datetime.now(timezone.utc).isoformat()
            await db.items.insert_one(new_item)

        created += 1

    return {"message": f"Item split into {created} pieces for order {req.order_no}"}

# ==========================================
# ADDONS
# ==========================================

@api_router.post("/addons")
async def add_addons(req: AddOnRequest):
    item = await db.items.find_one({"id": req.item_id}, {"_id": 0})
    if not item:
        raise HTTPException(status_code=404, detail="Item not found")

    addon_names = []
    total_amount = 0
    for addon in req.addons:
        addon_names.append(f"{addon['name']}({addon['price']})")
        total_amount += float(addon['price'])

    new_desc = ", ".join(addon_names)
    existing_desc = item.get("addon_desc", "N/A")
    if existing_desc and existing_desc != "N/A":
        new_desc = f"{existing_desc} + {new_desc}"

    new_total = float(item.get("addon_amount", 0)) + total_amount

    await db.items.update_one({"id": req.item_id}, {"$set": {
        "addon_desc": new_desc,
        "addon_amount": new_total,
        "addon_pay_mode": "Pending",
        "addon_pending": new_total,
    }})
    return {"message": "Add-ons saved", "addon_desc": new_desc, "addon_amount": new_total}

# ==========================================
# JOB WORK (Status Updates)
# ==========================================

@api_router.get("/jobwork")
async def get_jobwork(
    tab: str = "tailoring",
    order_no: Optional[str] = None,
    date_filter: Optional[str] = None,
    delivery_filter: Optional[str] = None,
):
    query = {}

    if tab == "tailoring":
        query["tailoring_status"] = {"$in": ["Pending", "Stitched", "Delivered"]}
    else:
        query["embroidery_status"] = {"$in": ["Required", "In Progress", "Finished"]}

    if order_no and order_no != "All":
        query["order_no"] = order_no
    if date_filter and date_filter != "All":
        query["date"] = date_filter
    if delivery_filter and delivery_filter != "All":
        query["delivery_date"] = delivery_filter

    items = await db.items.find(query, {"_id": 0}).sort("date", -1).to_list(500)

    if tab == "tailoring":
        pending = [i for i in items if i["tailoring_status"] == "Pending"]
        stitched = [i for i in items if i["tailoring_status"] == "Stitched"]
        delivered = [i for i in items if i["tailoring_status"] == "Delivered"]
        return {"pending": pending, "stitched": stitched, "delivered": delivered}
    else:
        required = [i for i in items if i["embroidery_status"] == "Required"]
        in_progress = [i for i in items if i["embroidery_status"] == "In Progress"]
        finished = [i for i in items if i["embroidery_status"] == "Finished"]
        return {"required": required, "in_progress": in_progress, "finished": finished}

@api_router.post("/jobwork/move")
async def move_jobwork(req: StatusUpdateRequest):
    updated = 0
    for item_id in req.item_ids:
        update_fields = {}
        if req.new_status in ["Pending", "Stitched", "Delivered"]:
            update_fields["tailoring_status"] = req.new_status
        elif req.new_status in ["Required", "In Progress", "Finished"]:
            update_fields["embroidery_status"] = req.new_status
            if req.karigar:
                update_fields["karigar"] = req.karigar

        result = await db.items.update_one({"id": item_id}, {"$set": update_fields})
        if result.modified_count > 0:
            updated += 1

    return {"message": f"{updated} items moved to {req.new_status}"}

class MoveBackRequest(BaseModel):
    item_ids: List[str]
    current_status: str

@api_router.post("/jobwork/move-back")
async def move_jobwork_back(req: MoveBackRequest):
    TAILORING_PREV = {"Stitched": "Pending", "Delivered": "Stitched"}
    EMB_PREV = {"In Progress": "Required", "Finished": "In Progress"}
    updated = 0
    for item_id in req.item_ids:
        if req.current_status in TAILORING_PREV:
            update_fields = {"tailoring_status": TAILORING_PREV[req.current_status]}
        elif req.current_status in EMB_PREV:
            update_fields = {"embroidery_status": EMB_PREV[req.current_status]}
        else:
            continue
        result = await db.items.update_one({"id": item_id}, {"$set": update_fields})
        if result.modified_count > 0:
            updated += 1
    return {"message": f"{updated} items moved back"}

class EmbMoveRequest(BaseModel):
    item_ids: List[str]
    new_status: str
    emb_labour_amount: Optional[float] = None
    emb_customer_amount: Optional[float] = None

@api_router.post("/jobwork/move-emb")
async def move_embroidery(req: EmbMoveRequest):
    updated = 0
    for item_id in req.item_ids:
        update_fields = {"embroidery_status": req.new_status}
        if req.emb_labour_amount is not None and req.emb_labour_amount > 0:
            update_fields["emb_labour_amount"] = req.emb_labour_amount
        if req.emb_customer_amount is not None and req.emb_customer_amount > 0:
            update_fields["embroidery_amount"] = req.emb_customer_amount
            update_fields["embroidery_pending"] = req.emb_customer_amount
            update_fields["embroidery_pay_mode"] = "Pending"
        result = await db.items.update_one({"id": item_id}, {"$set": update_fields})
        if result.modified_count > 0:
            updated += 1
    return {"message": f"{updated} embroidery items updated"}

class EmbEditRequest(BaseModel):
    item_id: str
    karigar: Optional[str] = None
    emb_labour_amount: Optional[float] = None
    emb_customer_amount: Optional[float] = None

@api_router.post("/jobwork/edit-emb")
async def edit_embroidery(req: EmbEditRequest):
    update_fields = {}
    if req.karigar is not None:
        update_fields["karigar"] = req.karigar
    if req.emb_labour_amount is not None:
        update_fields["emb_labour_amount"] = req.emb_labour_amount
    if req.emb_customer_amount is not None:
        update_fields["embroidery_amount"] = req.emb_customer_amount
        update_fields["embroidery_pending"] = req.emb_customer_amount
        update_fields["embroidery_pay_mode"] = "Pending"
    if not update_fields:
        return {"message": "Nothing to update"}
    result = await db.items.update_one({"id": req.item_id}, {"$set": update_fields})
    return {"message": "Updated" if result.modified_count > 0 else "No change"}

@api_router.get("/jobwork/filters")
async def get_jobwork_filters():
    order_nos = await db.items.distinct("order_no", {"order_no": {"$ne": "N/A"}})
    dates = await db.items.distinct("date")
    delivery_dates = await db.items.distinct("delivery_date", {"delivery_date": {"$ne": "N/A"}})
    return {
        "order_nos": sorted([o for o in order_nos if o]),
        "dates": sorted([d for d in dates if d], reverse=True),
        "delivery_dates": sorted([d for d in delivery_dates if d], reverse=True),
    }

# ==========================================
# SETTLEMENTS
# ==========================================

@api_router.get("/settlements/balances")
async def get_settlement_balances(name: Optional[str] = None, ref: Optional[str] = None):
    if not ref:
        return {"fabric": 0, "tailoring": 0, "embroidery": 0, "addon": 0, "advance": 0}

    not_settled = {"$not": {"$regex": "^Settled"}}
    pipeline_fab = [
        {"$match": {"ref": ref, "fabric_amount": {"$gt": 0}, "fabric_pay_mode": not_settled}},
        {"$group": {"_id": None, "total": {"$sum": "$fabric_pending"}}}
    ]
    pipeline_tail = [
        {"$match": {"ref": ref, "tailoring_amount": {"$gt": 0}, "tailoring_pay_mode": not_settled}},
        {"$group": {"_id": None, "total": {"$sum": "$tailoring_pending"}}}
    ]
    pipeline_emb = [
        {"$match": {"ref": ref, "embroidery_amount": {"$gt": 0}, "embroidery_pay_mode": not_settled}},
        {"$group": {"_id": None, "total": {"$sum": "$embroidery_pending"}}}
    ]
    pipeline_addon = [
        {"$match": {"ref": ref, "addon_amount": {"$gt": 0}, "addon_pay_mode": not_settled}},
        {"$group": {"_id": None, "total": {"$sum": "$addon_pending"}}}
    ]
    pipeline_adv = [
        {"$match": {"ref": ref}},
        {"$group": {"_id": None, "total": {"$sum": "$amount"}}}
    ]

    fab = await db.items.aggregate(pipeline_fab).to_list(1)
    tail = await db.items.aggregate(pipeline_tail).to_list(1)
    emb = await db.items.aggregate(pipeline_emb).to_list(1)
    addon = await db.items.aggregate(pipeline_addon).to_list(1)
    adv = await db.advances.aggregate(pipeline_adv).to_list(1)

    return {
        "fabric": max(0, fab[0]["total"]) if fab else 0,
        "tailoring": max(0, tail[0]["total"]) if tail else 0,
        "embroidery": max(0, emb[0]["total"]) if emb else 0,
        "addon": max(0, addon[0]["total"]) if addon else 0,
        "advance": adv[0]["total"] if adv else 0,
    }

@api_router.post("/settlements/pay")
async def process_settlement(req: SettlementRequest):
    modes_str = ", ".join(req.payment_modes) if req.payment_modes else "Cash"
    total_allocated = dq_round_money(
        req.allot_fabric + req.allot_tailoring + req.allot_embroidery + req.allot_addon + req.allot_advance
    )
    fresh_payment = dq_round_money(req.fresh_payment)

    if total_allocated <= 0:
        raise HTTPException(status_code=400, detail="Please allocate at least some amount")

    current_balances = await get_settlement_balances(ref=req.ref)

    available_advance = dq_round_money(current_balances["advance"])
    advance_to_use = 0.0
    if req.use_advance:
        advance_to_use = min(available_advance, max(0.0, dq_round_money(total_allocated - fresh_payment)))

    # Over-payment is allowed: excess is distributed pro-rata and pending goes negative.
    # Pool-match is validated on the frontend as a warning, not a hard block here.

    async def apply_pro_rata(ref, pay_mode_field, pay_date_field, received_field, pending_field, total_to_pay):
        # Derive the amount field name from the pending field name
        # e.g. "fabric_pending" -> "fabric_amount", "tailoring_pending" -> "tailoring_amount"
        amount_field = pending_field.replace("_pending", "_amount")

        # Fetch ALL items for this ref that have a non-zero category amount.
        # We cannot filter by pending>0 because over-payment must also reach
        # already-settled items (pending=0). Pro-rata weight = category amount.
        all_items = await db.items.find({"ref": ref}, {"_id": 0}).to_list(500)
        eligible = [i for i in all_items if dq_round_money(i.get(amount_field, 0)) > 0]
        if not eligible:
            return

        total_weight = sum(dq_round_money(i.get(amount_field, 0)) for i in eligible)
        running_paid = 0

        for idx, item in enumerate(eligible):
            weight = dq_round_money(item.get(amount_field, 0))
            bal = dq_round_money(item.get(pending_field, 0))
            if idx == len(eligible) - 1:
                share = dq_round_money(total_to_pay - running_paid)
            else:
                share = dq_round_money((weight / total_weight) * total_to_pay) if total_weight > 0 else 0
                running_paid += share

            existing_received = dq_round_money(item.get(received_field, 0))
            new_received = dq_round_money(existing_received + share)
            new_balance = dq_round_money(bal - share)  # negative when over-paid
            update = {
                pay_date_field: req.payment_date,
                received_field: new_received,
                pending_field: new_balance,
                pay_mode_field: dq_build_payment_mode_label(req.payment_modes, new_balance, new_received),
            }
            await db.items.update_one({"id": item["id"]}, {"$set": update})

    if req.allot_fabric > 0:
        await apply_pro_rata(req.ref, "fabric_pay_mode", "fabric_pay_date", "fabric_received", "fabric_pending", req.allot_fabric)
    if req.allot_tailoring > 0:
        await apply_pro_rata(req.ref, "tailoring_pay_mode", "tailoring_pay_date", "tailoring_received", "tailoring_pending", req.allot_tailoring)
    if req.allot_embroidery > 0:
        await apply_pro_rata(req.ref, "embroidery_pay_mode", "embroidery_pay_date", "embroidery_received", "embroidery_pending", req.allot_embroidery)
    if req.allot_addon > 0:
        await apply_pro_rata(req.ref, "addon_pay_mode", "addon_pay_date", "addon_received", "addon_pending", req.allot_addon)

    if req.allot_advance > 0:
        adv = {
            "id": str(uuid.uuid4()),
            "date": req.payment_date,
            "name": req.customer_name,
            "ref": req.ref,
            "amount": req.allot_advance,
            "mode": modes_str,
            "tally": False,
            "created_at": datetime.now(timezone.utc).isoformat(),
        }
        await db.advances.insert_one(adv)

    if advance_to_use > 0:
            adjustment = {
                "id": str(uuid.uuid4()),
                "date": req.payment_date,
                "name": req.customer_name,
                "ref": req.ref,
                "amount": -advance_to_use,
                "mode": "Adjusted",
                "tally": False,
                "created_at": datetime.now(timezone.utc).isoformat(),
            }
            await db.advances.insert_one(adjustment)

    return {"message": "Settlement processed successfully"}

# ==========================================
# DAYBOOK
# ==========================================

@api_router.get("/daybook")
async def get_daybook(date_filter: Optional[str] = None):
    # Key: (date, ref) — each unique pay-date × ref combination is a separate row
    entries = {}

    def get_or_create(date, ref, name):
        key = (date, ref)
        if key not in entries:
            entries[key] = {
                "date": date,
                "ref": ref,
                "name": name,
                "fabric": 0, "tailoring": 0, "embroidery": 0, "addon": 0, "advance": 0, "total": 0,
                "modes": {"fabric": "", "tailoring": "", "embroidery": "", "addon": "", "advance": ""},
                "tally_status": {"fabric": False, "tailoring": False, "embroidery": False, "addon": False, "advance": False},
            }
        return entries[key]

    item_query = {}
    if date_filter and date_filter != "All":
        item_query["$or"] = [
            {"fabric_pay_date": date_filter},
            {"tailoring_pay_date": date_filter},
            {"embroidery_pay_date": date_filter},
            {"addon_pay_date": date_filter},
        ]

    items = await db.items.find(item_query if date_filter and date_filter != "All" else {}, {"_id": 0}).to_list(2000)

    categories = [
        ("fabric",     "fabric_pay_date",     "fabric_received",     "fabric_pay_mode",     "tally_fabric"),
        ("tailoring",  "tailoring_pay_date",  "tailoring_received",  "tailoring_pay_mode",  "tally_tailoring"),
        ("embroidery", "embroidery_pay_date", "embroidery_received", "embroidery_pay_mode", "tally_embroidery"),
        ("addon",      "addon_pay_date",      "addon_received",      "addon_pay_mode",      "tally_addon"),
    ]

    for item in items:
        ref  = item.get("ref", "")
        name = item.get("name", "")
        for cat_name, date_field, received_field, mode_field, tally_field in categories:
            pay_date = item.get(date_field, "N/A")
            received = item.get(received_field, 0)
            if pay_date == "N/A" or not received:
                continue
            if date_filter and date_filter != "All" and pay_date != date_filter:
                continue

            e = get_or_create(pay_date, ref, name)
            e[cat_name]  += received
            e["total"]   += received
            e["tally_status"][cat_name] = item.get(tally_field, False)
            mode = item.get(mode_field, "")
            if mode:
                e["modes"][cat_name] = mode

    adv_query = {}
    if date_filter and date_filter != "All":
        adv_query["date"] = date_filter

    advances = await db.advances.find(adv_query, {"_id": 0}).to_list(500)
    for adv in advances:
        amount = adv.get("amount", 0)
        if not amount:
            continue
        ref      = adv.get("ref", "")
        adv_date = adv.get("date", "")
        if not adv_date:
            continue
        if date_filter and date_filter != "All" and adv_date != date_filter:
            continue

        e = get_or_create(adv_date, ref, adv.get("name", ""))
        e["advance"] += amount
        e["total"]   += amount
        e["tally_status"]["advance"] = adv.get("tally", False)
        mode = adv.get("mode", "")
        if mode:
            e["modes"]["advance"] = mode

    return {"entries": list(entries.values())}

@api_router.get("/daybook/dates")
async def get_daybook_dates():
    dates = set()
    for field in ["fabric_pay_date", "tailoring_pay_date", "embroidery_pay_date", "addon_pay_date"]:
        vals = await db.items.distinct(field)
        for v in vals:
            if v and v != "N/A":
                dates.add(v)

    adv_dates = await db.advances.distinct("date")
    for v in adv_dates:
        if v:
            dates.add(v)

    return sorted(list(dates), reverse=True)

@api_router.post("/daybook/tally")
async def tally_entries(req: TallyRequest):
    tally_value = req.action == "tally"

    date_field_map = {
        "fabric":     ("tally_fabric",     "fabric_pay_date"),
        "tailoring":  ("tally_tailoring",  "tailoring_pay_date"),
        "embroidery": ("tally_embroidery", "embroidery_pay_date"),
        "addon":      ("tally_addon",      "addon_pay_date"),
    }

    if req.category == "advance":
        for entry_ref in req.entry_ids:
            adv_query = {"ref": entry_ref}
            if req.date:
                adv_query["date"] = req.date
            await db.advances.update_many(adv_query, {"$set": {"tally": tally_value}})
    elif req.category == "all":
        for entry_ref in req.entry_ids:
            # Update each category scoped to its own pay_date
            for cat, (tally_field, pay_date_field) in date_field_map.items():
                item_query = {"ref": entry_ref}
                if req.date:
                    item_query[pay_date_field] = req.date
                await db.items.update_many(item_query, {"$set": {tally_field: tally_value}})
            adv_query = {"ref": entry_ref}
            if req.date:
                adv_query["date"] = req.date
            await db.advances.update_many(adv_query, {"$set": {"tally": tally_value}})
    elif req.category in date_field_map:
        tally_field, pay_date_field = date_field_map[req.category]
        for entry_ref in req.entry_ids:
            item_query = {"ref": entry_ref}
            if req.date:
                item_query[pay_date_field] = req.date
            await db.items.update_many(item_query, {"$set": {tally_field: tally_value}})

    return {"message": f"{len(req.entry_ids)} entries {req.action}ed"}

# ==========================================
# LABOUR PAYMENTS
# ==========================================

@api_router.get("/labour")
async def get_labour_items(filter_type: str = "All", filter_karigar: str = "All", view_mode: str = "unpaid"):
    items = []
    paid = view_mode == "paid"

    if filter_type in ("All", "Tailoring Labour"):
        if paid:
            query = {
                "tailoring_status": {"$in": ["Stitched", "Delivered"]},
                "labour_paid": "Yes",
                "labour_amount": {"$gt": 0},
            }
        else:
            query = {
                "tailoring_status": {"$in": ["Stitched", "Delivered"]},
                "labour_paid": {"$in": ["N/A", "", None]},
                "labour_amount": {"$gt": 0},
            }
        tail_items = await db.items.find(query, {"_id": 0}).to_list(500)
        for item in tail_items:
            karigar = item.get("karigar", "N/A")
            if filter_karigar == "All" or karigar == filter_karigar:
                items.append({**item, "labour_type": "Tailoring"})

    if filter_type in ("All", "Embroidery Labour"):
        if paid:
            query = {
                "embroidery_status": "Finished",
                "emb_labour_paid": "Yes",
                "emb_labour_amount": {"$gt": 0},
            }
        else:
            query = {
                "embroidery_status": "Finished",
                "emb_labour_paid": {"$in": ["N/A", "", None]},
                "emb_labour_amount": {"$gt": 0},
            }
        emb_items = await db.items.find(query, {"_id": 0}).to_list(500)
        for item in emb_items:
            karigar = item.get("karigar", "N/A")
            if filter_karigar == "All" or karigar == filter_karigar:
                items.append({**item, "labour_type": "Embroidery"})

    return items

@api_router.get("/labour/karigars")
async def get_karigars():
    karigars = await db.items.distinct("karigar", {"karigar": {"$nin": ["N/A", "", None]}})
    return sorted(karigars)

@api_router.post("/labour/pay")
async def pay_labour(req: LabourPaymentRequest):
    updated = 0
    mode_str = ", ".join(req.payment_modes) if req.payment_modes else "Cash"
    for item_id in req.item_ids:
        if req.labour_type == "tailoring":
            update = {
                "labour_paid": "Yes",
                "labour_pay_date": req.payment_date,
                "labour_payment_mode": mode_str,
                "labour_payment_id": req.payment_id or "",
            }
        else:
            update = {
                "emb_labour_paid": "Yes",
                "emb_labour_date": req.payment_date,
                "emb_labour_payment_mode": mode_str,
                "emb_labour_payment_id": req.payment_id or "",
            }
        result = await db.items.update_one({"id": item_id}, {"$set": update})
        if result.modified_count > 0:
            updated += 1

    return {"message": f"{updated} labour payments processed"}

class LabourDeleteRequest(BaseModel):
    payment_id: Optional[str] = None
    item_ids: List[str]
    labour_type: str

@api_router.post("/labour/delete-payment")
async def delete_labour_payment(req: LabourDeleteRequest):
    updated = 0
    for item_id in req.item_ids:
        if req.labour_type == "tailoring":
            update = {"labour_paid": "N/A", "labour_pay_date": "N/A"}
        else:
            update = {"emb_labour_paid": "N/A", "emb_labour_date": "N/A"}
        result = await db.items.update_one({"id": item_id}, {"$set": update})
        if result.modified_count > 0:
            updated += 1
    return {"message": f"{updated} items marked as unpaid"}

# ==========================================
# ADVANCES
# ==========================================

@api_router.get("/advances")
async def get_advances(name: Optional[str] = None, ref: Optional[str] = None):
    query = {}
    if name:
        query["name"] = name
    if ref:
        query["ref"] = ref
    advances = await db.advances.find(query, {"_id": 0}).sort("date", -1).to_list(500)
    return advances

class AdvanceCreateRequest(BaseModel):
    ref: str
    name: str
    amount: float
    date: str
    mode: Optional[str] = "Cash"

@api_router.post("/advances")
async def create_advance(req: AdvanceCreateRequest):
    import uuid
    adv = {"id": str(uuid.uuid4()), "ref": req.ref, "name": req.name, "amount": req.amount, "date": req.date, "mode": req.mode or "Cash"}
    await db.advances.insert_one(adv)
    adv.pop("_id", None)
    return adv

@api_router.put("/advances/{advance_id}")
async def update_advance(advance_id: str, req: AdvanceCreateRequest):
    update = {"ref": req.ref, "name": req.name, "amount": req.amount, "date": req.date, "mode": req.mode or "Cash"}
    result = await db.advances.update_one({"id": advance_id}, {"$set": update})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Advance not found")
    return {"message": "Advance updated"}

@api_router.delete("/advances/{advance_id}")
async def delete_advance(advance_id: str):
    result = await db.advances.delete_one({"id": advance_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Advance not found")
    return {"message": "Advance deleted"}

# ==========================================
# ORDER NUMBERS
# ==========================================

@api_router.get("/orders")
async def get_order_numbers():
    orders = await db.items.distinct("order_no", {"order_no": {"$nin": ["N/A", "", None]}})
    return sorted([o for o in orders if o])

@api_router.get("/orders/status")
async def get_order_status(
    customer: Optional[str] = None,
    order_no: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    limit: int = Query(400, le=2000)
):
    query = {"order_no": {"$nin": ["N/A", "", None]}}
    if customer:
        query["name"] = customer
    if order_no:
        query["order_no"] = {"$regex": order_no, "$options": "i"}
    if date_from:
        query.setdefault("date", {})["$gte"] = date_from
    if date_to:
        query.setdefault("date", {})["$lte"] = date_to
    items = await db.items.find(query, {"_id": 0}).to_list(limit)
    grouped: dict = {}
    for item in items:
        ono = item.get("order_no", "")
        if ono not in grouped:
            grouped[ono] = {
                "order_no": ono,
                "customers": set(),
                "refs": set(),
                "item_count": 0,
                "tailoring_pending": 0,
                "tailoring_stitched": 0,
                "tailoring_delivered": 0,
                "emb_required": 0,
                "emb_in_progress": 0,
                "emb_finished": 0,
                "order_total": 0,
                "latest_bill_date": "",
                "latest_delivery_date": "",
            }
        g = grouped[ono]
        g["customers"].add(item.get("name", ""))
        g["refs"].add(item.get("ref", ""))
        g["item_count"] += 1
        ts = item.get("tailoring_status", "N/A")
        if ts == "Pending": g["tailoring_pending"] += 1
        elif ts == "Stitched": g["tailoring_stitched"] += 1
        elif ts == "Delivered": g["tailoring_delivered"] += 1
        es = item.get("embroidery_status", "N/A")
        if es == "Required": g["emb_required"] += 1
        elif es == "In Progress": g["emb_in_progress"] += 1
        elif es == "Finished": g["emb_finished"] += 1
        g["order_total"] += float(item.get("fabric_amount", 0))
        d = item.get("date", "")
        if d and d > g["latest_bill_date"]: g["latest_bill_date"] = d
        dd = item.get("delivery_date", "")
        if dd and dd not in ("N/A", "", None) and dd > g["latest_delivery_date"]: g["latest_delivery_date"] = dd
    result = []
    for g in grouped.values():
        g["customers"] = sorted(g["customers"])
        g["refs"] = sorted(g["refs"])
        result.append(g)
    result.sort(key=lambda x: x.get("latest_bill_date", ""), reverse=True)
    return result

# ==========================================
# ITEM EDIT & DELETE
# ==========================================

@api_router.put("/items/{item_id}")
async def update_item(item_id: str, req: ItemUpdateRequest):
    item = await db.items.find_one({"id": item_id})
    if not item:
        raise HTTPException(status_code=404, detail="Item not found")

    update_fields = {}
    for field, value in req.model_dump(exclude_unset=True).items():
        if value is not None:
            update_fields[field] = value

    # Recalculate fabric_amount if price/qty/discount changed
    if any(f in update_fields for f in ["price", "qty", "discount"]):
        p = update_fields.get("price", item.get("price", 0))
        q = update_fields.get("qty", item.get("qty", 0))
        d = update_fields.get("discount", item.get("discount", 0))
        update_fields["fabric_amount"] = round((p - (p * d / 100)) * q, 0)

    if update_fields:
        await db.items.update_one({"id": item_id}, {"$set": update_fields})

    updated = await db.items.find_one({"id": item_id}, {"_id": 0})
    return updated

@api_router.delete("/items/{item_id}")
async def delete_item(item_id: str):
    result = await db.items.delete_one({"id": item_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Item not found")
    return {"message": "Item deleted"}

@api_router.delete("/items/bulk/delete")
async def bulk_delete_items(item_ids: List[str]):
    result = await db.items.delete_many({"id": {"$in": item_ids}})
    return {"message": f"{result.deleted_count} items deleted"}

# ==========================================
# SEARCH
# ==========================================

@api_router.get("/search")
async def search_items(
    q: str = "",
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    customer: Optional[str] = None,
    status: Optional[str] = None,
    payment_status: Optional[str] = None,
    min_amount: Optional[float] = None,
    max_amount: Optional[float] = None,
    limit: int = 100,
    skip: int = 0,
):
    query = {}

    if q:
        query["$or"] = [
            {"name": {"$regex": q, "$options": "i"}},
            {"barcode": {"$regex": q, "$options": "i"}},
            {"ref": {"$regex": q, "$options": "i"}},
            {"article_type": {"$regex": q, "$options": "i"}},
            {"order_no": {"$regex": q, "$options": "i"}},
            {"karigar": {"$regex": q, "$options": "i"}},
            {"addon_desc": {"$regex": q, "$options": "i"}},
        ]

    if customer and customer != "All":
        query["name"] = customer

    if date_from:
        query.setdefault("date", {})["$gte"] = date_from
    if date_to:
        query.setdefault("date", {})["$lte"] = date_to

    if status and status != "All":
        if status in ["Pending", "Stitched", "Delivered", "Awaiting Order", "N/A"]:
            query["tailoring_status"] = status
        elif status in ["Required", "In Progress", "Finished", "Not Required"]:
            query["embroidery_status"] = status

    if payment_status and payment_status != "All":
        if payment_status == "Settled":
            query["fabric_pay_mode"] = {"$regex": "^Settled"}
        elif payment_status == "Pending":
            query["fabric_pay_mode"] = {"$not": {"$regex": "^Settled"}}
            query["fabric_amount"] = {"$gt": 0}

    if min_amount is not None:
        query.setdefault("fabric_amount", {})["$gte"] = min_amount
    if max_amount is not None:
        query.setdefault("fabric_amount", {})["$lte"] = max_amount

    items = await db.items.find(query, {"_id": 0}).sort("date", -1).skip(skip).limit(limit).to_list(limit)
    for item in items:
        item["payment_status"] = dq_determine_payment_status(item.get("fabric_pending", 0), item.get("fabric_received", 0))
    total = await db.items.count_documents(query)
    return {"items": items, "total": total}

# ==========================================
# HTML INVOICE (print-ready, loads in iframe)
# ==========================================

@api_router.get("/invoice")
async def generate_invoice(ref_id: str = Query(..., alias="ref")):
    from fastapi.responses import HTMLResponse

    items = await db.items.find({"ref": ref_id}, {"_id": 0}).to_list(100)
    if not items:
        raise HTTPException(status_code=404, detail="No items found for this reference")

    advances = await db.advances.find({"ref": ref_id}, {"_id": 0}).to_list(50)
    stored_settings = await db.settings.find_one({"key": "app_settings"}, {"_id": 0})
    s = merge_settings(stored_settings)

    GST_RATE = float(s.get("gst_rate", DEFAULT_SETTINGS["gst_rate"]))
    brand_color = s.get("firm_name_color", "#C86B4D")
    firm_name = s.get("firm_name", DEFAULT_SETTINGS["firm_name"])
    firm_name_case = s.get("firm_name_case", "uppercase")
    firm_name_size = s.get("firm_name_size", "16")
    firm_address = s.get("firm_address", DEFAULT_SETTINGS["firm_address"])
    firm_phones = s.get("firm_phones", DEFAULT_SETTINGS["firm_phones"])
    firm_gstin = s.get("firm_gstin", DEFAULT_SETTINGS["firm_gstin"])
    firm_logo = s.get("firm_logo", "")

    customer_name = items[0].get("name", "N/A")
    order_date = items[0].get("date", "N/A")

    def fmt(n):
        try:
            return f"{float(n):,.0f}"
        except:
            return "0"

    def calc_gst(amt):
        base = round(float(amt) / (1 + GST_RATE / 100), 2)
        gst = round(float(amt) - base, 2)
        return base, gst

    # ---- Fabric rows ----
    fab_total = 0
    fab_gst_total = 0
    fab_rows_html = ""
    for i, item in enumerate(items, 1):
        amt = float(item.get("fabric_amount", 0))
        fab_total += amt
        base, gst = calc_gst(amt)
        fab_gst_total += gst
        art = item.get("article_type", "") or "-"
        fab_rows_html += f"""
        <tr>
          <td>{i}</td><td>{art}</td>
          <td>{str(item.get("barcode",""))[:20]}</td>
          <td class="r">₹{fmt(item.get("price",0))}</td>
          <td class="r">{item.get("qty",0)}</td>
          <td class="r">{float(item.get("discount",0)):.0f}%</td>
          <td class="r">₹{fmt(amt)}</td>
          <td class="r">₹{base:,.2f}</td>
          <td class="r">₹{gst:,.2f}</td>
        </tr>"""
    fab_base_total = round(fab_total - fab_gst_total, 2)
    fab_foot = f"""<tr class="foot">
      <td colspan="6" class="r">TOTAL</td>
      <td class="r">₹{fmt(fab_total)}</td>
      <td class="r">₹{fab_base_total:,.2f}</td>
      <td class="r">₹{fab_gst_total:,.2f}</td>
    </tr>"""

    # ---- Tailoring section ----
    tail_items = [x for x in items if x.get("tailoring_status") not in ("N/A", None, "", "Awaiting Order") and float(x.get("tailoring_amount", 0)) > 0]
    tail_section = ""
    tail_total = 0
    if tail_items:
        rows = ""
        for i, ti in enumerate(tail_items, 1):
            t_amt = float(ti.get("tailoring_amount", 0))
            tail_total += t_amt
            pm = ti.get("tailoring_pay_mode", "N/A")
            status_lbl = "Settled" if str(pm).startswith("Settled") else pm
            rows += f"""<tr>
              <td>{i}</td><td>{ti.get("article_type","")}</td><td>{ti.get("order_no","")}</td>
              <td>{ti.get("delivery_date","")}</td><td class="r">₹{fmt(t_amt)}</td>
              <td>{status_lbl}</td>
            </tr>"""
        rows += f'<tr class="foot"><td colspan="4" class="r">TOTAL</td><td class="r">₹{fmt(tail_total)}</td><td></td></tr>'
        tail_section = f"""
        <h3 class="sec-title">Tailoring</h3>
        <table><thead><tr><th>#</th><th>Article</th><th>Order#</th><th>Delivery</th><th class="r">Tailoring</th><th>Payment</th></tr></thead>
        <tbody>{rows}</tbody></table>"""

    # ---- Embroidery section ----
    emb_items = [x for x in items if x.get("embroidery_status") not in ("N/A", "Not Required", None, "")]
    emb_section = ""
    emb_total = 0
    if emb_items:
        rows = ""
        for i, ei in enumerate(emb_items, 1):
            e_amt = float(ei.get("embroidery_amount", 0))
            emb_total += e_amt
            pm = ei.get("embroidery_pay_mode", "N/A")
            emb_status = ei.get("embroidery_status", "")
            amt_display = f'₹{fmt(e_amt)}' if e_amt > 0 else '<em style="color:#9C9690">To be Calculated</em>'
            rows += f"""<tr>
              <td>{i}</td><td>{ei.get("article_type","")}</td><td>{emb_status}</td>
              <td>{ei.get("karigar","N/A")}</td><td class="r">{amt_display}</td>
              <td>{"Settled" if str(pm).startswith("Settled") else pm}</td>
            </tr>"""
        emb_total_display = f'₹{fmt(emb_total)}' if emb_total > 0 else '<em style="color:#9C9690">To be Calculated</em>'
        rows += f'<tr class="foot"><td colspan="4" class="r">TOTAL</td><td class="r">{emb_total_display}</td><td></td></tr>'
        emb_section = f"""
        <h3 class="sec-title">Embroidery</h3>
        <table><thead><tr><th>#</th><th>Article</th><th>Status</th><th>Karigar</th><th class="r">Amount</th><th>Payment</th></tr></thead>
        <tbody>{rows}</tbody></table>"""

    # ---- Add-ons section ----
    addon_items = [x for x in items if x.get("addon_desc") not in ("N/A", None, "")]
    addon_section = ""
    addon_total = 0
    if addon_items:
        rows = ""
        for i, ai in enumerate(addon_items, 1):
            a_amt = float(ai.get("addon_amount", 0))
            addon_total += a_amt
            pm = ai.get("addon_pay_mode", "N/A")
            rows += f"""<tr>
              <td>{i}</td><td>{ai.get("article_type","")}</td><td>{ai.get("addon_desc","")}</td>
              <td class="r">₹{fmt(a_amt)}</td>
              <td>{"Settled" if str(pm).startswith("Settled") else pm}</td>
            </tr>"""
        rows += f'<tr class="foot"><td colspan="3" class="r">TOTAL</td><td class="r">₹{fmt(addon_total)}</td><td></td></tr>'
        addon_section = f"""
        <h3 class="sec-title">Add-ons</h3>
        <table><thead><tr><th>#</th><th>Article</th><th>Add-on</th><th class="r">Amount</th><th>Payment</th></tr></thead>
        <tbody>{rows}</tbody></table>"""

    # ---- Advances section ----
    total_adv = sum(float(a.get("amount", 0)) for a in advances)
    adv_section = ""
    if total_adv != 0:
        rows = "".join(f'<tr><td>{a.get("date","")}</td><td class="r">₹{fmt(a.get("amount",0))}</td><td>{a.get("mode","")}</td></tr>' for a in advances)
        rows += f'<tr class="foot"><td>Net Advance</td><td class="r">₹{fmt(total_adv)}</td><td></td></tr>'
        adv_section = f"""
        <h3 class="sec-title">Advances</h3>
        <table><thead><tr><th>Date</th><th class="r">Amount</th><th>Mode</th></tr></thead>
        <tbody>{rows}</tbody></table>"""

    # ---- Payment summary ----
    fab_received = sum(float(i.get("fabric_received", 0)) for i in items)
    fab_pending = sum(float(i.get("fabric_pending", 0)) for i in items if not str(i.get("fabric_pay_mode", "")).startswith("Settled"))
    tail_total_amt = sum(float(i.get("tailoring_amount", 0)) for i in items)
    tail_received = sum(float(i.get("tailoring_received", 0)) for i in items)
    tail_pending_amt = sum(float(i.get("tailoring_pending", 0)) for i in items if not str(i.get("tailoring_pay_mode", "")).startswith("Settled"))
    emb_total_amt = sum(float(i.get("embroidery_amount", 0)) for i in items)
    emb_received = sum(float(i.get("embroidery_received", 0)) for i in items)
    emb_pending_amt = sum(float(i.get("embroidery_pending", 0)) for i in items if not str(i.get("embroidery_pay_mode", "")).startswith("Settled"))
    addon_total_amt = sum(float(i.get("addon_amount", 0)) for i in items)
    addon_received = sum(float(i.get("addon_received", 0)) for i in items)
    addon_pending_amt = sum(float(i.get("addon_pending", 0)) for i in items if not str(i.get("addon_pay_mode", "")).startswith("Settled"))

    grand_total = fab_total + tail_total_amt + emb_total_amt + addon_total_amt
    total_received = fab_received + tail_received + emb_received + addon_received
    total_pending = fab_pending + tail_pending_amt + emb_pending_amt + addon_pending_amt
    net_payable = max(total_pending - max(total_adv, 0), 0)

    summary_rows = f"""
      <tr><td>Fabric (incl. GST {GST_RATE:.0f}%)</td><td class="r">₹{fmt(fab_total)}</td><td class="r">₹{fmt(fab_received)}</td><td class="r">₹{fmt(fab_pending)}</td></tr>"""
    if tail_total_amt > 0:
        summary_rows += f'<tr><td>Tailoring</td><td class="r">₹{fmt(tail_total_amt)}</td><td class="r">₹{fmt(tail_received)}</td><td class="r">₹{fmt(tail_pending_amt)}</td></tr>'
    if emb_total_amt > 0:
        summary_rows += f'<tr><td>Embroidery</td><td class="r">₹{fmt(emb_total_amt)}</td><td class="r">₹{fmt(emb_received)}</td><td class="r">₹{fmt(emb_pending_amt)}</td></tr>'
    if addon_total_amt > 0:
        summary_rows += f'<tr><td>Add-ons</td><td class="r">₹{fmt(addon_total_amt)}</td><td class="r">₹{fmt(addon_received)}</td><td class="r">₹{fmt(addon_pending_amt)}</td></tr>'
    summary_rows += f"""
      <tr class="foot grand"><td>GRAND TOTAL</td><td class="r">₹{fmt(grand_total)}</td><td class="r">₹{fmt(total_received)}</td><td class="r">₹{fmt(total_pending)}</td></tr>"""
    if total_adv > 0:
        summary_rows += f'<tr><td>Less: Advance</td><td></td><td class="r">₹{fmt(total_adv)}</td><td></td></tr>'
        summary_rows += f'<tr class="foot net"><td colspan="3" class="r">NET PAYABLE</td><td class="r">₹{fmt(net_payable)}</td></tr>'

    logo_html = ""
    if firm_logo:
        logo_url = firm_logo if firm_logo.startswith("http") else f"{firm_logo}"
        logo_html = f'<img src="{logo_url}" alt="logo" class="logo" />'

    gen_time = datetime.now().strftime("%d-%m-%Y %H:%M")

    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8"/>
<meta name="viewport" content="width=device-width,initial-scale=1"/>
<title>Invoice – {ref_id}</title>
<style>
  @import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700&display=swap');
  *, *::before, *::after {{ box-sizing: border-box; margin: 0; padding: 0; }}
  body {{ font-family: 'Inter', Arial, sans-serif; font-size: 12px; color: #2D2A26; background: #fff; padding: 28px 32px; }}
  .invoice-wrap {{ max-width: 860px; margin: 0 auto; }}

  /* ── Centered firm header ── */
  .header {{ text-align: center; padding-bottom: 16px; border-bottom: 2.5px solid {brand_color}; margin-bottom: 16px; }}
  .logo {{ width: 72px; height: 72px; object-fit: contain; border-radius: 6px; display: block; margin: 0 auto 8px; }}
  .firm-name {{ font-size: {firm_name_size}pt; font-weight: 700; color: {brand_color}; text-transform: {firm_name_case}; letter-spacing: 0.05em; }}
  .firm-sub {{ font-size: 10.5px; color: #6C6760; margin-top: 3px; line-height: 1.5; }}
  .invoice-badge {{ display: inline-block; margin-top: 10px; padding: 3px 16px; background: {brand_color}; color: #fff; font-size: 11px; font-weight: 700; letter-spacing: 0.18em; text-transform: uppercase; border-radius: 2px; }}

  /* ── Bill meta strip ── */
  .meta-strip {{ display: grid; grid-template-columns: repeat(4, 1fr); gap: 0; margin-bottom: 18px; border: 1px solid #E0DDD7; border-radius: 4px; overflow: hidden; }}
  .meta-cell {{ padding: 8px 12px; border-right: 1px solid #E0DDD7; }}
  .meta-cell:last-child {{ border-right: none; }}
  .meta-label {{ font-size: 9.5px; text-transform: uppercase; letter-spacing: 0.1em; color: #9C9690; margin-bottom: 2px; }}
  .meta-value {{ font-size: 12px; font-weight: 600; color: #2D2A26; }}

  /* ── Section titles ── */
  .sec-title {{ font-size: 10.5px; font-weight: 700; text-transform: uppercase; letter-spacing: 0.1em; color: {brand_color}; margin: 20px 0 5px; padding-bottom: 4px; border-bottom: 1.5px solid #E0DDD7; }}

  /* ── Tables ── */
  table {{ width: 100%; border-collapse: collapse; font-size: 11px; margin-bottom: 4px; }}
  thead tr {{ background: #3D3A36; }}
  th {{ padding: 6px 8px; font-weight: 600; text-align: left; color: #F5F3EE; white-space: nowrap; font-size: 10.5px; letter-spacing: 0.03em; }}
  td {{ padding: 5px 8px; border-bottom: 1px solid #EDEBE6; vertical-align: top; }}
  tbody tr:nth-child(even) {{ background: #FAFAF8; }}
  tr:last-child td {{ border-bottom: none; }}
  tr.foot td {{ font-weight: 700; background: #EDEBE6; border-top: 1.5px solid #3D3A36; border-bottom: none; }}
  .r {{ text-align: right; }}

  /* ── Payment summary ── */
  .summary-wrap {{ margin-top: 22px; }}
  tr.grand td {{ font-size: 12.5px; font-weight: 700; background: #3D3A36; color: #F5F3EE; border-top: none; border-bottom: none; }}
  tr.grand td.r {{ text-align: right; }}
  tr.net td {{ font-size: 13.5px; font-weight: 700; color: #fff; background: {brand_color}; border-top: none; border-bottom: none; }}

  /* ── Terms ── */
  .terms {{ margin-top: 22px; padding-top: 12px; border-top: 1px dashed #CBCAC4; display: grid; grid-template-columns: 1fr 1fr; gap: 2px 24px; }}
  .terms p {{ font-size: 10px; color: #8C8A84; margin-bottom: 2px; }}

  /* ── Footer ── */
  .inv-footer {{ margin-top: 16px; font-size: 10px; color: #AEACA6; text-align: center; padding-top: 10px; border-top: 1px solid #EDEBE6; }}

  /* ── Print ── */
  @media print {{
    body {{ padding: 0; -webkit-print-color-adjust: exact; print-color-adjust: exact; color: #1A1A1A; }}
    .no-print {{ display: none !important; }}
    thead tr {{ background: #3D3A36 !important; -webkit-print-color-adjust: exact; print-color-adjust: exact; }}
    th {{ color: #F5F3EE !important; }}
    tr.foot td {{ background: #EDEBE6 !important; }}
    tr.grand td {{ background: #3D3A36 !important; color: #F5F3EE !important; }}
    tr.net td {{ background: {brand_color} !important; color: #fff !important; }}
    .meta-strip {{ border: 1px solid #AAAAAA !important; -webkit-print-color-adjust: exact; print-color-adjust: exact; }}
    .meta-cell {{ border-right: 1px solid #AAAAAA !important; }}
    .meta-label {{ color: #444444 !important; }}
    .meta-value {{ color: #1A1A1A !important; }}
    .header {{ border-bottom: 2.5px solid {brand_color} !important; }}
    .firm-sub {{ color: #444444 !important; }}
    .sec-title {{ color: {brand_color} !important; border-bottom-color: #AAAAAA !important; }}
    .terms p {{ color: #444444 !important; }}
    .inv-footer {{ color: #555555 !important; border-top-color: #AAAAAA !important; }}
    td em {{ color: #555555 !important; }}
    @page {{ margin: 10mm 12mm; size: A4; }}
  }}
</style>
</head>
<body>
<div class="invoice-wrap">

  <!-- CENTERED FIRM HEADER -->
  <div class="header">
    {logo_html}
    <div class="firm-name">{firm_name}</div>
    <div class="firm-sub">{firm_address}</div>
    <div class="firm-sub">Ph: {firm_phones} &nbsp;&nbsp;|&nbsp;&nbsp; GSTIN: {firm_gstin}</div>
    <div><span class="invoice-badge">Invoice</span></div>
  </div>

  <!-- BILL META STRIP -->
  <div class="meta-strip">
    <div class="meta-cell"><div class="meta-label">Customer</div><div class="meta-value">{customer_name}</div></div>
    <div class="meta-cell"><div class="meta-label">Reference</div><div class="meta-value">{ref_id}</div></div>
    <div class="meta-cell"><div class="meta-label">Date</div><div class="meta-value">{order_date}</div></div>
    <div class="meta-cell"><div class="meta-label">Items</div><div class="meta-value">{len(items)}</div></div>
  </div>

  <!-- A. FABRIC ITEMS -->
  <h3 class="sec-title">Fabric Items</h3>
  <table>
    <thead><tr><th>#</th><th>Article</th><th>Barcode</th><th class="r">Price/m</th><th class="r">Qty</th><th class="r">Disc%</th><th class="r">Amount</th><th class="r">Base Amt</th><th class="r">GST ({GST_RATE:.0f}%)</th></tr></thead>
    <tbody>{fab_rows_html}{fab_foot}</tbody>
  </table>

  {tail_section}
  {emb_section}
  {addon_section}
  {adv_section}

  <!-- F. PAYMENT SUMMARY -->
  <div class="summary-wrap">
    <h3 class="sec-title">Payment Summary</h3>
    <table>
      <thead><tr><th>Category</th><th class="r">Total</th><th class="r">Received</th><th class="r">Pending</th></tr></thead>
      <tbody>{summary_rows}</tbody>
    </table>
  </div>

  <!-- TERMS -->
  <div class="terms">
    <p>1. Fabric prices are inclusive of GST @ {GST_RATE:.0f}%.</p>
    <p>2. Goods once sold will not be taken back or exchanged.</p>
    <p>3. Tailoring subject to delivery timelines agreed at order.</p>
    <p>4. Advances are non-refundable and adjusted against final bill.</p>
    <p>5. Any dispute is subject to local jurisdiction.</p>
  </div>

  <div class="inv-footer">Invoice Ref: <strong>{ref_id}</strong> &nbsp;·&nbsp; Generated: {gen_time} &nbsp;·&nbsp; Thank you for your business!</div>
</div>
</body>
</html>"""

    return HTMLResponse(content=html, status_code=200)

# ==========================================
# REPORTS & ANALYTICS
# ==========================================

@api_router.get("/reports/revenue")
async def get_revenue_report(period: str = "daily", date_from: Optional[str] = None, date_to: Optional[str] = None):
    match_query = {}
    if date_from:
        match_query.setdefault("date", {})["$gte"] = date_from
    if date_to:
        match_query.setdefault("date", {})["$lte"] = date_to

    pipeline = [
        {"$match": match_query} if match_query else {"$match": {}},
        {"$group": {
            "_id": "$date",
            "fabric_total": {"$sum": "$fabric_amount"},
            "fabric_received": {"$sum": "$fabric_received"},
            "tailoring_total": {"$sum": "$tailoring_amount"},
            "tailoring_received": {"$sum": "$tailoring_received"},
            "embroidery_total": {"$sum": "$embroidery_amount"},
            "embroidery_received": {"$sum": "$embroidery_received"},
            "addon_total": {"$sum": "$addon_amount"},
            "count": {"$sum": 1},
        }},
        {"$sort": {"_id": 1}},
    ]

    daily = await db.items.aggregate(pipeline).to_list(1000)

    if period == "weekly":
        weekly = {}
        for d in daily:
            try:
                dt = datetime.strptime(d["_id"], "%Y-%m-%d")
                week_start = dt.strftime("%Y-W%W")
                if week_start not in weekly:
                    weekly[week_start] = {"_id": week_start, "fabric_total": 0, "fabric_received": 0, "tailoring_total": 0, "tailoring_received": 0, "embroidery_total": 0, "embroidery_received": 0, "addon_total": 0, "count": 0}
                for k in ["fabric_total", "fabric_received", "tailoring_total", "tailoring_received", "embroidery_total", "embroidery_received", "addon_total", "count"]:
                    weekly[week_start][k] += d[k]
            except Exception:
                pass
        return list(weekly.values())

    if period == "monthly":
        monthly = {}
        for d in daily:
            month_key = d["_id"][:7] if d["_id"] else "unknown"
            if month_key not in monthly:
                monthly[month_key] = {"_id": month_key, "fabric_total": 0, "fabric_received": 0, "tailoring_total": 0, "tailoring_received": 0, "embroidery_total": 0, "embroidery_received": 0, "addon_total": 0, "count": 0}
            for k in ["fabric_total", "fabric_received", "tailoring_total", "tailoring_received", "embroidery_total", "embroidery_received", "addon_total", "count"]:
                monthly[month_key][k] += d[k]
        return list(monthly.values())

    return daily

@api_router.get("/reports/customers")
async def get_customer_report():
    pipeline = [
        {"$group": {
            "_id": "$name",
            "total_fabric": {"$sum": "$fabric_amount"},
            "total_received": {"$sum": "$fabric_received"},
            "total_pending_raw": {"$sum": {"$cond": [{"$not": [{"$regexMatch": {"input": {"$ifNull": ["$fabric_pay_mode", ""]}, "regex": "^Settled"}}]}, "$fabric_pending", 0]}},
            "total_tailoring": {"$sum": "$tailoring_amount"},
            "items_count": {"$sum": 1},
            "refs": {"$addToSet": "$ref"},
        }},
        {"$sort": {"total_fabric": -1}},
    ]
    result = await db.items.aggregate(pipeline).to_list(200)
    return [
        {
            "name": r["_id"],
            "total_fabric": r["total_fabric"],
            "total_received": r["total_received"],
            "total_pending": max(0, r["total_pending_raw"]),
            "total_tailoring": r["total_tailoring"],
            "items_count": r["items_count"],
            "refs_count": len(r["refs"]),
        }
        for r in result if r["_id"]
    ]

@api_router.get("/reports/summary")
async def get_summary_report(date_from: Optional[str] = None, date_to: Optional[str] = None):
    match_query = {}
    if date_from:
        match_query.setdefault("date", {})["$gte"] = date_from
    if date_to:
        match_query.setdefault("date", {})["$lte"] = date_to

    items = await db.items.find(match_query if match_query else {}, {"_id": 0}).to_list(5000)
    advances = await db.advances.find({}, {"_id": 0}).to_list(500)

    total_fabric = sum(i.get("fabric_amount", 0) for i in items)
    total_fabric_received = sum(i.get("fabric_received", 0) for i in items)
    # max(0, sum) so over-paid credits (negative pending) correctly reduce the outstanding total
    total_fabric_pending = sum(i.get("fabric_pending", 0) for i in items if not str(i.get("fabric_pay_mode", "")).startswith("Settled"))
    total_tailoring = sum(i.get("tailoring_amount", 0) for i in items)
    total_tailoring_received = sum(i.get("tailoring_received", 0) for i in items)
    total_tailoring_pending = sum(i.get("tailoring_pending", 0) for i in items if not str(i.get("tailoring_pay_mode", "")).startswith("Settled"))
    total_embroidery = sum(i.get("embroidery_amount", 0) for i in items)
    total_embroidery_received = sum(i.get("embroidery_received", 0) for i in items)
    total_embroidery_pending = sum(i.get("embroidery_pending", 0) for i in items if not str(i.get("embroidery_pay_mode", "")).startswith("Settled"))
    total_addon = sum(i.get("addon_amount", 0) for i in items)
    total_addon_pending = sum(i.get("addon_pending", 0) for i in items if not str(i.get("addon_pay_mode", "")).startswith("Settled"))
    total_advance = sum(a.get("amount", 0) for a in advances)

    # Payment mode breakdown
    mode_counts = {}
    for i in items:
        mode = i.get("fabric_pay_mode", "N/A")
        if mode.startswith("Settled"):
            parts = mode.replace("Settled - ", "").split(", ")
            for p in parts:
                p = p.strip()
                if p:
                    mode_counts[p] = mode_counts.get(p, 0) + i.get("fabric_received", 0)

    # Article type breakdown
    article_counts = {}
    for i in items:
        at = i.get("article_type", "N/A")
        if at != "N/A":
            article_counts[at] = article_counts.get(at, 0) + 1

    return {
        "total_fabric": total_fabric,
        "total_fabric_received": total_fabric_received,
        "total_fabric_pending": total_fabric_pending,
        "total_tailoring": total_tailoring,
        "total_tailoring_received": total_tailoring_received,
        "total_tailoring_pending": total_tailoring_pending,
        "total_embroidery": total_embroidery,
        "total_embroidery_received": total_embroidery_received,
        "total_embroidery_pending": total_embroidery_pending,
        "total_addon": total_addon,
        "total_addon_pending": total_addon_pending,
        "total_advance": total_advance,
        "total_items": len(items),
        "payment_modes": [{"mode": k, "amount": v} for k, v in sorted(mode_counts.items(), key=lambda x: -x[1])],
        "article_types": [{"type": k, "count": v} for k, v in sorted(article_counts.items(), key=lambda x: -x[1])],
    }

# ==========================================
# EXCEL IMPORT (Upload .xlsm/.xlsx from browser)
# ==========================================

@api_router.post("/import/excel")
async def import_excel(file: UploadFile = File(...), mode: str = "replace"):
    if not file.filename.endswith(('.xlsm', '.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Please upload an Excel file (.xlsm or .xlsx)")

    try:
        import openpyxl
        contents = await file.read()
        wb = openpyxl.load_workbook(io.BytesIO(contents), data_only=True)

        items_count = 0
        advances_count = 0
        items = []
        advances = []

        if 'Item Details' in wb.sheetnames:
            ws = wb['Item Details']
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
                if not row[0]:
                    continue

                def safe_float(v):
                    if v is None or str(v).strip() in ("N/A", "", "None"):
                        return 0
                    try:
                        return float(v)
                    except (ValueError, TypeError):
                        return 0

                def safe_str(v):
                    if v is None:
                        return "N/A"
                    return str(v).strip()

                def safe_date(v):
                    if v and hasattr(v, 'strftime'):
                        return v.strftime("%Y-%m-%d")
                    return "N/A"

                item = {
                    "id": str(uuid.uuid4()),
                    "date": safe_date(row[0]),
                    "name": safe_str(row[1]),
                    "ref": safe_str(row[2]),
                    "barcode": safe_str(row[3]),
                    "price": safe_float(row[4]),
                    "qty": safe_float(row[5]),
                    "discount": safe_float(row[6]),
                    "fabric_amount": safe_float(row[7]),
                    "tailoring_status": safe_str(row[8]),
                    "article_type": safe_str(row[9]),
                    "order_no": safe_str(row[10]),
                    "delivery_date": safe_date(row[11]),
                    "tailoring_amount": safe_float(row[12]),
                    "embroidery_status": safe_str(row[13]),
                    "embroidery_amount": safe_float(row[14]),
                    "addon_desc": safe_str(row[15]),
                    "addon_amount": safe_float(row[16]),
                    "fabric_pay_mode": safe_str(row[17]),
                    "fabric_pay_date": safe_date(row[18]),
                    "fabric_pending": safe_float(row[19]),
                    "fabric_received": safe_float(row[20]),
                    "labour_amount": safe_float(row[21]),
                    "labour_paid": safe_str(row[22]),
                    "labour_pay_date": safe_date(row[23]),
                    "tailoring_pay_mode": safe_str(row[24]),
                    "tailoring_pay_date": safe_date(row[25]),
                    "tailoring_received": safe_float(row[26]),
                    "tailoring_pending": safe_float(row[27]),
                    "embroidery_pay_mode": safe_str(row[28]),
                    "embroidery_pay_date": safe_date(row[29]),
                    "embroidery_received": safe_float(row[30]),
                    "embroidery_pending": safe_float(row[31]),
                    "addon_pay_mode": safe_str(row[32]),
                    "addon_pay_date": safe_date(row[33]),
                    "addon_received": safe_float(row[34]),
                    "addon_pending": safe_float(row[35]),
                    "karigar": safe_str(row[36]) if len(row) > 36 else "N/A",
                    "tally_fabric": False,
                    "tally_tailoring": False,
                    "tally_embroidery": False,
                    "tally_addon": False,
                    "created_at": datetime.now(timezone.utc).isoformat(),
                }
                items.append(item)

            if items:
                items_count = len(items)

        if 'Advances' in wb.sheetnames:
            ws2 = wb['Advances']
            for row in ws2.iter_rows(min_row=2, max_row=ws2.max_row, values_only=True):
                if not row[0]:
                    continue
                dt = row[0]
                date_str = dt.strftime("%Y-%m-%d") if hasattr(dt, 'strftime') else str(dt)[:10]
                adv = {
                    "id": str(uuid.uuid4()),
                    "date": date_str,
                    "name": str(row[1]).strip() if row[1] else "",
                    "ref": str(row[2]).strip() if row[2] else "",
                    "amount": float(row[3]) if row[3] else 0,
                    "mode": str(row[4]).strip() if row[4] else "",
                    "tally": False,
                    "created_at": datetime.now(timezone.utc).isoformat(),
                }
                advances.append(adv)

            if advances:
                advances_count = len(advances)

        if mode == "replace":
            await db.items.delete_many({})
            await db.advances.delete_many({})

        if items:
            await db.items.insert_many(items)

        if advances:
            await db.advances.insert_many(advances)

        return {
            "message": f"Import successful! {items_count} items and {advances_count} advances imported.",
            "items_count": items_count,
            "advances_count": advances_count,
        }
    except Exception as e:
        logger.error(f"Import error: {e}")
        raise HTTPException(status_code=500, detail=f"Import failed: {str(e)}")

# ==========================================
# EXCEL EXPORT
# ==========================================

@api_router.get("/export/excel")
async def export_excel():
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()

    # Items sheet
    ws = wb.active
    ws.title = "Item Details"
    headers = [
        "Date", "Name", "Ref.", "Items", "Price", "Qty", "Discount", "Fabric Amount",
        "Tailoring?", "Article Type", "Order No.", "Delivery Date", "Tailoring Amount",
        "Embroidery?", "Embroidery Amount", "Add-on", "Add-on Amount",
        "Fabric Payment Mode", "Fabric Payment Date", "Fabric Pending Balance", "Fabric Payment Received",
        "Labour Amount", "Labour Paid?", "Labour Payment Date",
        "Tailoring Payment Mode", "Tailoring Payment Date", "Tailoring Payment Received", "Tailoring Pending Balance",
        "Embroidery Payment Mode", "Embroidery Payment Date", "Embroidery Payment Received", "Embroidery Pending Balance",
        "Add-On Payment Mode", "Add-On Payment Date", "Add-On Payment Received", "Add-On Pending Balance", "Karigar?"
    ]

    header_fill = PatternFill(start_color="C86B4D", end_color="C86B4D", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=10)

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    items = await db.items.find({}, {"_id": 0}).sort("date", 1).to_list(10000)
    fields = [
        "date", "name", "ref", "barcode", "price", "qty", "discount", "fabric_amount",
        "tailoring_status", "article_type", "order_no", "delivery_date", "tailoring_amount",
        "embroidery_status", "embroidery_amount", "addon_desc", "addon_amount",
        "fabric_pay_mode", "fabric_pay_date", "fabric_pending", "fabric_received",
        "labour_amount", "labour_paid", "labour_pay_date",
        "tailoring_pay_mode", "tailoring_pay_date", "tailoring_received", "tailoring_pending",
        "embroidery_pay_mode", "embroidery_pay_date", "embroidery_received", "embroidery_pending",
        "addon_pay_mode", "addon_pay_date", "addon_received", "addon_pending", "karigar"
    ]

    for row_idx, item in enumerate(items, 2):
        for col_idx, field in enumerate(fields, 1):
            ws.cell(row=row_idx, column=col_idx, value=item.get(field, ""))

    # Advances sheet
    ws2 = wb.create_sheet("Advances")
    adv_headers = ["Advance Payment Date", "Name", "Ref", "Advance Payment Amount", "Advance Payment Mode"]
    for col, header in enumerate(adv_headers, 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font

    advances = await db.advances.find({}, {"_id": 0}).sort("date", 1).to_list(5000)
    for row_idx, adv in enumerate(advances, 2):
        ws2.cell(row=row_idx, column=1, value=adv.get("date", ""))
        ws2.cell(row=row_idx, column=2, value=adv.get("name", ""))
        ws2.cell(row=row_idx, column=3, value=adv.get("ref", ""))
        ws2.cell(row=row_idx, column=4, value=adv.get("amount", 0))
        ws2.cell(row=row_idx, column=5, value=adv.get("mode", ""))

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"retail_book_export_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

# ==========================================
# DATABASE BACKUP & RESTORE
# ==========================================

@api_router.get("/backup")
async def backup_database():
    items = await db.items.find({}, {"_id": 0}).to_list(50000)
    advances = await db.advances.find({}, {"_id": 0}).to_list(10000)

    backup_data = {
        "version": "1.0",
        "created_at": datetime.now(timezone.utc).isoformat(),
        "items_count": len(items),
        "advances_count": len(advances),
        "items": items,
        "advances": advances,
    }

    buffer = io.BytesIO(json.dumps(backup_data, indent=2, default=str).encode('utf-8'))
    buffer.seek(0)

    filename = f"retail_backup_{datetime.now().strftime('%Y%m%d_%H%M')}.json"
    return StreamingResponse(
        buffer,
        media_type="application/json",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@api_router.post("/restore")
async def restore_database(file: UploadFile = File(...)):
    if not file.filename.endswith('.json'):
        raise HTTPException(status_code=400, detail="Please upload a .json backup file")

    try:
        contents = await file.read()
        backup_data = json.loads(contents.decode('utf-8'))

        if "items" not in backup_data or "advances" not in backup_data:
            raise HTTPException(status_code=400, detail="Invalid backup file format")

        items_count = 0
        advances_count = 0
        items = backup_data["items"]
        advances = backup_data["advances"]

        if not isinstance(items, list) or not isinstance(advances, list):
            raise HTTPException(status_code=400, detail="Invalid backup file format")

        await db.items.delete_many({})
        await db.advances.delete_many({})

        if items:
            await db.items.insert_many(items)
            items_count = len(items)

        if advances:
            await db.advances.insert_many(advances)
            advances_count = len(advances)

        return {
            "message": f"Restore successful! {items_count} items and {advances_count} advances restored.",
            "items_count": items_count,
            "advances_count": advances_count,
        }
    except json.JSONDecodeError:
        raise HTTPException(status_code=400, detail="Invalid JSON file")
    except Exception as e:
        logger.error(f"Restore error: {e}")
        raise HTTPException(status_code=500, detail=f"Restore failed: {str(e)}")

@api_router.get("/db/stats")
async def get_db_stats():
    items_count = await db.items.count_documents({})
    advances_count = await db.advances.count_documents({})
    return {
        "items_count": items_count,
        "advances_count": advances_count,
        "db_name": os.environ.get('DB_NAME', 'unknown'),
    }

@api_router.get("/db/audit")
async def get_db_audit(limit: int = 100):
    safe_limit = max(1, min(limit, 500))
    return await dq_generate_data_audit(db, safe_limit)

@api_router.post("/db/normalize")
async def normalize_db_data(limit: int = 100):
    safe_limit = max(1, min(limit, 500))
    return await dq_normalize_low_risk_data(db, safe_limit)

@api_router.post("/db/repair")
async def repair_db_data(limit: int = 100):
    safe_limit = max(1, min(limit, 500))
    return await dq_repair_high_risk_data(db, safe_limit)

# ==========================================
# SETTINGS (authenticated)
# ==========================================

DEFAULT_SETTINGS = {
    "article_types": ARTICLE_TYPES,
    "tailoring_rates": {k: {"tailoring": v[0], "labour": v[1]} for k, v in TAILORING_RATES.items()},
    "payment_modes": PAYMENT_MODES,
    "addon_items": ADDON_ITEMS,
    "gst_rate": 5.0,
    "firm_name": "Narwana Agencies",
    "firm_address": "Jasmeet Nagar, Near Kalka Chowk, Ambala City, Pin: 134003, Haryana",
    "firm_phones": "9467902343, 7056212655",
    "firm_gstin": "06ADMPG9353K1Z4",
    "firm_logo": None,
    "firm_name_color": "#C86B4D",
    "firm_name_size": "16",
    "firm_name_case": "uppercase",
}

def merge_settings(stored_settings: Optional[dict] = None) -> dict:
    merged = dict(DEFAULT_SETTINGS)
    if stored_settings:
        merged.update({k: v for k, v in stored_settings.items() if k != "key"})
    return merged

@api_router.get("/settings/public")
async def get_public_settings():
    settings = await db.settings.find_one({"key": "app_settings"}, {"_id": 0})
    merged = merge_settings(settings)
    return {"firm_name": merged.get("firm_name", "Retail Book")}

@api_router.get("/settings")
async def get_settings(current_user: dict = Depends(get_current_user_dep)):
    settings = await db.settings.find_one({"key": "app_settings"}, {"_id": 0})
    return merge_settings(settings)

@api_router.put("/settings")
async def update_settings(data: dict, current_user: dict = Depends(get_current_user_dep)):
    if current_user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Only admins can update settings")
    data["key"] = "app_settings"
    await db.settings.update_one({"key": "app_settings"}, {"$set": data}, upsert=True)
    settings = await db.settings.find_one({"key": "app_settings"}, {"_id": 0})
    return merge_settings(settings)

# ==========================================
# LOGO UPLOAD
# ==========================================

@api_router.post("/upload/logo")
async def upload_logo(file: UploadFile = File(...)):
    if not file.content_type or not file.content_type.startswith("image/"):
        raise HTTPException(status_code=400, detail="Only image files allowed")
    if file.size and file.size > 1024 * 1024:
        raise HTTPException(status_code=400, detail="Image too large (max 1MB)")
    upload_dir = ROOT_DIR / "static" / "uploads"
    upload_dir.mkdir(parents=True, exist_ok=True)
    ext = file.filename.split(".")[-1].lower() if "." in file.filename else "png"
    safe_name = f"logo_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{os.urandom(4).hex()}.{ext}"
    file_path = upload_dir / safe_name
    contents = await file.read()
    with open(file_path, "wb") as f:
        f.write(contents)
    return {"url": f"/uploads/{safe_name}"}

# ==========================================
# AUTH ENDPOINTS
# ==========================================

@api_router.post("/auth/login")
async def login(req: LoginRequest, request: Request):
    client_ip = request.client.host if request.client else "unknown"
    _check_rate_limit(client_ip)
    user = await db.users.find_one({"username": req.username.lower().strip()})
    if not user:
        raise HTTPException(status_code=401, detail="Invalid username or password")
    if not auth_module.verify_password(req.password, user.get("password_hash", "")):
        raise HTTPException(status_code=401, detail="Invalid username or password")
    if not user.get("is_active", True):
        raise HTTPException(status_code=403, detail="User is disabled")
    _clear_rate_limit(client_ip)
    token = auth_module.create_access_token({"sub": user["username"]})
    return {
        "access_token": token,
        "token_type": "bearer",
        "user": {
            "username": user["username"],
            "full_name": user.get("full_name", ""),
            "role": user.get("role", "cashier"),
            "is_active": user.get("is_active", True),
            "allowed_pages": user.get("allowed_pages", []),
        }
    }

@api_router.post("/auth/logout")
async def logout(current_user: dict = Depends(get_current_user_dep)):
    return {"message": "Logged out"}

@api_router.get("/auth/me")
async def get_me(current_user: dict = Depends(get_current_user_dep)):
    return {
        "username": current_user["username"],
        "full_name": current_user.get("full_name", ""),
        "role": current_user.get("role", "cashier"),
        "is_active": current_user.get("is_active", True),
        "allowed_pages": current_user.get("allowed_pages", []),
    }

@api_router.post("/auth/register")
async def register_user(req: UserCreateRequest, current_user: dict = Depends(get_current_user_dep)):
    if current_user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Only admins can create users")
    existing = await db.users.find_one({"username": req.username.lower().strip()})
    if existing:
        raise HTTPException(status_code=400, detail="Username already exists")
    new_user = {
        "username": req.username,
        "password_hash": auth_module.get_password_hash(req.password),
        "full_name": req.full_name,
        "role": req.role,
        "is_active": True,
        "allowed_pages": req.allowed_pages,
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    new_user["username"] = req.username.lower().strip()
    await db.users.insert_one(new_user)
    logger.info(f"User '{new_user['username']}' created by '{current_user['username']}'")
    return {"message": "User created successfully", "username": new_user["username"]}

@api_router.get("/auth/users")
async def list_users(current_user: dict = Depends(get_current_user_dep)):
    if current_user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Only admins can list users")
    users = await db.users.find({}, {"password_hash": 0}).to_list(None)
    for u in users:
        u["_id"] = str(u["_id"])
    return users

@api_router.put("/auth/users/{username}")
async def update_user(username: str, data: dict, current_user: dict = Depends(get_current_user_dep)):
    if current_user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Only admins can update users")
    if username == "admin" and current_user["username"] != "admin":
        raise HTTPException(status_code=403, detail="Cannot modify the admin account")
    update = {}
    if "full_name" in data: update["full_name"] = data["full_name"]
    if "role" in data: update["role"] = data["role"]
    if "is_active" in data: update["is_active"] = data["is_active"]
    if "allowed_pages" in data: update["allowed_pages"] = data["allowed_pages"]
    if "password" in data and data["password"]:
        update["password_hash"] = auth_module.get_password_hash(data["password"])
    if not update:
        raise HTTPException(status_code=400, detail="No fields to update")
    result = await db.users.update_one({"username": username}, {"$set": update})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="User not found")
    logger.info(f"User '{username}' updated by '{current_user['username']}'")
    return {"message": "User updated successfully"}

@api_router.delete("/auth/users/{username}")
async def delete_user(username: str, current_user: dict = Depends(get_current_user_dep)):
    if current_user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Only admins can delete users")
    if username == current_user["username"]:
        raise HTTPException(status_code=400, detail="Cannot delete your own account")
    if username == "admin":
        raise HTTPException(status_code=400, detail="Cannot delete the admin account")
    result = await db.users.delete_one({"username": username})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="User not found")
    logger.info(f"User '{username}' deleted by '{current_user['username']}'")
    return {"message": "User deleted successfully"}

# ==========================================
# AUDIT LOGS
# ==========================================

@api_router.get("/audit-logs")
async def list_audit_logs(
    limit: int = Query(50, ge=1, le=500),
    skip: int = Query(0, ge=0),
    current_user: dict = Depends(get_current_user_dep),
):
    if current_user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Only admins can view audit logs")
    cursor = db.audit_logs.find().sort("timestamp", -1).skip(skip).limit(limit)
    docs = await cursor.to_list(length=limit)
    for d in docs:
        d["_id"] = str(d["_id"])
    return {"logs": docs, "count": len(docs)}

# ==========================================
# APP SETUP
# ==========================================

app.include_router(api_router)

app.add_middleware(GZipMiddleware, minimum_size=500)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

uploads_dir = ROOT_DIR / "static" / "uploads"
uploads_dir.mkdir(parents=True, exist_ok=True)
app.mount("/uploads", StaticFiles(directory=uploads_dir), name="uploads")

build_dir = ROOT_DIR / "frontend" / "build"
if build_dir.exists():
    app.mount("/static", StaticFiles(directory=build_dir / "static"), name="react-static")

    @app.get("/{full_path:path}", include_in_schema=False)
    async def spa_fallback(full_path: str):
        index = build_dir / "index.html"
        return FileResponse(str(index))

@app.on_event("startup")
async def startup_db_client():
    await db.items.create_index("id", unique=True, background=True)
    await db.items.create_index("ref", background=True)
    await db.items.create_index("barcode", background=True)
    await db.items.create_index("name", background=True)
    await db.items.create_index("date", background=True)
    await db.items.create_index("order_no", background=True)
    await db.items.create_index("karigar", background=True)
    await db.advances.create_index("id", unique=True, background=True)
    await db.advances.create_index("ref", background=True)
    await db.advances.create_index("date", background=True)
    await db.settings.create_index("key", unique=True, background=True)
    logger.info("MongoDB indexes ensured.")

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()
