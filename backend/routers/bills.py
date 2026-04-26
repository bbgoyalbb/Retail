"""
Bills router.
"""
from fastapi import APIRouter, HTTPException, Query, Depends, Request
from typing import Optional, List, Dict, Any
from datetime import datetime, timezone, date
import uuid
import re
import os
import logging
from bson import ObjectId
from .deps import db, get_current_user_dep
from data_quality import round_money, determine_payment_status, build_payment_mode_label
import auth as auth_module
from auth import audit_log
from .models import ADDON_ITEMS, ARTICLE_TYPES, BillLineItem, CreateBillRequest, PAYMENT_MODES, TAILORING_RATES, validate_date

logger = logging.getLogger(__name__)

router = APIRouter()

@router.post("/seed")
async def seed_data(current_user: dict = Depends(get_current_user_dep)):
    count = await db.items.count_documents({})
    if count > 0:
        return {"message": "Data already seeded", "items_count": count}

    try:
        import openpyxl
        wb_path = "/tmp/retail_book.xlsm"
        if not os.path.exists(wb_path):
            return {"message": "Excel file not found at /tmp/retail_book.xlsm"}

        wb = openpyxl.load_workbook(wb_path, data_only=True)

        def safe_float(v):
            if v is None or str(v).strip() in ("N/A", "", "None"):
                return 0
            try:
                return float(v)
            except (ValueError, TypeError):
                return 0

        def safe_str(v):
            if v is None:
                return "N/A"
            return str(v).strip()

        ws = wb['Item Details']
        items = []
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
            if not row[0]:
                continue
            dt = row[0]
            date_str = dt.strftime("%Y-%m-%d") if hasattr(dt, 'strftime') else str(dt)[:10]

            fab_pay_date = ""
            if row[18] and hasattr(row[18], 'strftime'):
                fab_pay_date = row[18].strftime("%Y-%m-%d")

            delivery_date = ""
            if row[11] and hasattr(row[11], 'strftime'):
                delivery_date = row[11].strftime("%Y-%m-%d")

            tail_pay_date = ""
            if row[25] and hasattr(row[25], 'strftime'):
                tail_pay_date = row[25].strftime("%Y-%m-%d")

            emb_pay_date = ""
            if row[29] and hasattr(row[29], 'strftime'):
                emb_pay_date = row[29].strftime("%Y-%m-%d")

            addon_pay_date = ""
            if row[33] and hasattr(row[33], 'strftime'):
                addon_pay_date = row[33].strftime("%Y-%m-%d")

            labour_pay_date = ""
            if row[23] and hasattr(row[23], 'strftime'):
                labour_pay_date = row[23].strftime("%Y-%m-%d")

            item = {
                "id": str(uuid.uuid4()),
                "date": date_str,
                "name": safe_str(row[1]),
                "ref": safe_str(row[2]),
                "barcode": safe_str(row[3]),
                "price": safe_float(row[4]),
                "qty": safe_float(row[5]),
                "discount": safe_float(row[6]),
                "fabric_amount": safe_float(row[7]),
                "tailoring_status": safe_str(row[8]),
                "article_type": safe_str(row[9]),
                "order_no": safe_str(row[10]),
                "delivery_date": delivery_date if delivery_date else "N/A",
                "tailoring_amount": safe_float(row[12]),
                "embroidery_status": safe_str(row[13]),
                "embroidery_amount": safe_float(row[14]),
                "addon_desc": safe_str(row[15]),
                "addon_amount": safe_float(row[16]),
                "fabric_pay_mode": safe_str(row[17]),
                "fabric_pay_date": fab_pay_date if fab_pay_date else "N/A",
                "fabric_pending": safe_float(row[19]),
                "fabric_received": safe_float(row[20]),
                "labour_amount": safe_float(row[21]),
                "labour_paid": safe_str(row[22]),
                "labour_pay_date": labour_pay_date if labour_pay_date else "N/A",
                "tailoring_pay_mode": safe_str(row[24]),
                "tailoring_pay_date": tail_pay_date if tail_pay_date else "N/A",
                "tailoring_received": safe_float(row[26]),
                "tailoring_pending": safe_float(row[27]),
                "embroidery_pay_mode": safe_str(row[28]),
                "embroidery_pay_date": emb_pay_date if emb_pay_date else "N/A",
                "embroidery_received": safe_float(row[30]),
                "embroidery_pending": safe_float(row[31]),
                "addon_pay_mode": safe_str(row[32]),
                "addon_pay_date": addon_pay_date if addon_pay_date else "N/A",
                "addon_received": safe_float(row[34]),
                "addon_pending": safe_float(row[35]),
                "karigar": safe_str(row[36]) if len(row) > 36 else "N/A",
                "tally_fabric": False,
                "tally_tailoring": False,
                "tally_embroidery": False,
                "tally_addon": False,
                "created_at": datetime.now(timezone.utc).isoformat(),
            }
            items.append(item)

        if items:
            await db.items.insert_many(items)

        advances = []
        if 'Advances' in wb.sheetnames:
            ws2 = wb['Advances']
            for row in ws2.iter_rows(min_row=2, max_row=ws2.max_row, values_only=True):
                if not row[0]:
                    continue
                dt = row[0]
                date_str = dt.strftime("%Y-%m-%d") if hasattr(dt, 'strftime') else str(dt)[:10]
                adv = {
                    "id": str(uuid.uuid4()),
                    "date": date_str,
                    "name": str(row[1]).strip() if row[1] else "",
                    "ref": str(row[2]).strip() if row[2] else "",
                    "amount": float(row[3]) if row[3] else 0,
                    "mode": str(row[4]).strip() if row[4] else "",
                    "tally": False,
                    "created_at": datetime.now(timezone.utc).isoformat(),
                }
                advances.append(adv)

        if advances:
            await db.advances.insert_many(advances)

        return {"message": "Data seeded successfully", "items_count": len(items), "advances_count": len(advances)}
    except Exception as e:
        logger.error(f"Seed error: {e}")
        raise HTTPException(status_code=500, detail=str(e))

# ==========================================
# DASHBOARD
# ==========================================

@router.get("/dashboard")
async def get_dashboard(current_user: dict = Depends(get_current_user_dep)):
    total_items = await db.items.count_documents({})
    total_advances = await db.advances.count_documents({})

    _not_settled = {"$not": {"$regex": "^Settled"}}
    pipeline_fabric_pending = [
        {"$match": {"fabric_amount": {"$gt": 0}, "fabric_pay_mode": _not_settled}},
        {"$group": {"_id": None, "total": {"$sum": "$fabric_pending"}}}
    ]
    fab_pending = await db.items.aggregate(pipeline_fabric_pending).to_list(1)

    pipeline_tail_pending = [
        {"$match": {"tailoring_amount": {"$gt": 0}, "tailoring_pay_mode": _not_settled}},
        {"$group": {"_id": None, "total": {"$sum": "$tailoring_pending"}}}
    ]
    tail_pending = await db.items.aggregate(pipeline_tail_pending).to_list(1)

    pipeline_emb_pending = [
        {"$match": {"embroidery_amount": {"$gt": 0}, "embroidery_pay_mode": _not_settled}},
        {"$group": {"_id": None, "total": {"$sum": "$embroidery_pending"}}}
    ]
    emb_pending = await db.items.aggregate(pipeline_emb_pending).to_list(1)

    pipeline_addon_pending = [
        {"$match": {"addon_amount": {"$gt": 0}, "addon_pay_mode": _not_settled}},
        {"$group": {"_id": None, "total": {"$sum": "$addon_pending"}}}
    ]
    addon_pending = await db.items.aggregate(pipeline_addon_pending).to_list(1)

    tailoring_pending_count = await db.items.count_documents({"tailoring_status": "Pending"})
    tailoring_stitched_count = await db.items.count_documents({"tailoring_status": "Stitched"})
    emb_required = await db.items.count_documents({"embroidery_status": "Required"})
    emb_inprogress = await db.items.count_documents({"embroidery_status": "In Progress"})

    unique_customers = await db.items.distinct("name")

    pipeline_revenue = [
        {"$group": {"_id": None, "total": {"$sum": "$fabric_received"}}}
    ]
    revenue = await db.items.aggregate(pipeline_revenue).to_list(1)

    pipeline_adv_total = [
        {"$group": {"_id": None, "total": {"$sum": "$amount"}}}
    ]
    adv_total = await db.advances.aggregate(pipeline_adv_total).to_list(1)

    # Get recent unique bill refs (grouped by ref) with customer info and totals
    pipeline_recent = [
        {"$sort": {"date": -1, "ref": -1}},
        {"$group": {
            "_id": "$ref",
            "date": {"$first": "$date"},
            "name": {"$first": "$name"},
            "fabric_total": {"$sum": "$fabric_amount"},
            "item_count": {"$sum": 1}
        }},
        {"$sort": {"date": -1}},
        {"$limit": 10},
        {"$project": {
            "_id": 0,
            "ref": "$_id",
            "date": 1,
            "name": 1,
            "fabric_total": 1,
            "item_count": 1
        }}
    ]
    recent_items = await db.items.aggregate(pipeline_recent).to_list(10)

    # 7-day revenue sparkline trend
    from datetime import date, timedelta
    trend_data = []
    for i in range(6, -1, -1):
        day = (date.today() - timedelta(days=i)).isoformat()
        day_pipeline = [
            {"$match": {"date": day}},
            {"$group": {"_id": None, "total": {"$sum": "$fabric_received"}}}
        ]
        day_result = await db.items.aggregate(day_pipeline).to_list(1)
        trend_data.append(day_result[0]["total"] if day_result else 0)

    return {
        "total_items": total_items,
        "revenue_trend": trend_data,
        "total_advances": total_advances,
        "fabric_pending_amount": fab_pending[0]["total"] if fab_pending else 0,
        "tailoring_pending_amount": tail_pending[0]["total"] if tail_pending else 0,
        "embroidery_pending_amount": emb_pending[0]["total"] if emb_pending else 0,
        "addon_pending_amount": addon_pending[0]["total"] if addon_pending else 0,
        "tailoring_pending_count": tailoring_pending_count,
        "tailoring_stitched_count": tailoring_stitched_count,
        "embroidery_required_count": emb_required,
        "embroidery_inprogress_count": emb_inprogress,
        "unique_customers": len(unique_customers),
        "total_revenue": revenue[0]["total"] if revenue else 0,
        "total_advances_amount": adv_total[0]["total"] if adv_total else 0,
        "recent_items": recent_items,
    }

# ==========================================
# CUSTOMERS
# ==========================================

@router.get("/customers")
async def get_customers(pending_only: bool = False, current_user: dict = Depends(get_current_user_dep)):
    if pending_only:
        _ns = {"$not": {"$regex": "^Settled"}}
        pipeline = [
            {"$match": {"$or": [
                {"fabric_amount": {"$gt": 0}, "fabric_pay_mode": _ns},
                {"tailoring_amount": {"$gt": 0}, "tailoring_pay_mode": _ns},
                {"embroidery_amount": {"$gt": 0}, "embroidery_pay_mode": _ns},
                {"addon_amount": {"$gt": 0}, "addon_pay_mode": _ns},
            ]}},
            {"$group": {"_id": "$name"}},
        ]
        result = await db.items.aggregate(pipeline).to_list(1000)
        return sorted([r["_id"] for r in result if r["_id"] and r["_id"] != "N/A"])
    customers = await db.items.distinct("name")
    return sorted([c for c in customers if c and c != "N/A"])

# ==========================================
# ITEMS CRUD
# ==========================================

@router.get("/items")
async def get_items(
    name: Optional[str] = None,
    ref: Optional[str] = None,
    date: Optional[str] = None,
    tailoring_status: Optional[str] = None,
    embroidery_status: Optional[str] = None,
    order_no: Optional[str] = None,
    limit: int = 500,
    skip: int = 0,
    current_user: dict = Depends(get_current_user_dep),
):
    query = {}
    if name:
        query["name"] = name
    if ref:
        query["ref"] = ref
    if date:
        query["date"] = date
    if tailoring_status:
        query["tailoring_status"] = tailoring_status
    if embroidery_status:
        query["embroidery_status"] = embroidery_status
    if order_no:
        query["order_no"] = order_no

    items = await db.items.find(query, {"_id": 0}).sort("date", -1).skip(skip).limit(limit).to_list(limit)
    total = await db.items.count_documents(query)
    return {"items": items, "total": total}

@router.get("/items/{item_id}")
async def get_item(item_id: str, current_user: dict = Depends(get_current_user_dep)):
    item = await db.items.find_one({"id": item_id}, {"_id": 0})
    if not item:
        raise HTTPException(status_code=404, detail="Item not found")
    return item

@router.get("/refs")
async def get_refs(name: Optional[str] = None, pending_only: bool = False, current_user: dict = Depends(get_current_user_dep)):
    query = {}
    if name:
        query["name"] = name
    if pending_only:
        _ns = {"$not": {"$regex": "^Settled"}}
        query["$or"] = [
            {"fabric_amount": {"$gt": 0}, "fabric_pay_mode": _ns},
            {"tailoring_amount": {"$gt": 0}, "tailoring_pay_mode": _ns},
            {"embroidery_amount": {"$gt": 0}, "embroidery_pay_mode": _ns},
            {"addon_amount": {"$gt": 0}, "addon_pay_mode": _ns},
        ]
    pipeline = [
        {"$match": query},
        {"$group": {"_id": "$ref"}},
        {"$sort": {"_id": -1}}
    ]
    refs = await db.items.aggregate(pipeline).to_list(500)
    return [r["_id"] for r in refs if r["_id"] and r["_id"] != "N/A"]

# ==========================================
# NEW BILL
# ==========================================

@router.post("/bills")
async def create_bill(req: CreateBillRequest, current_user: dict = Depends(get_current_user_dep)):
    if not req.items:
        raise HTTPException(status_code=400, detail="At least one item is required")

    # Validate dates
    try:
        validate_date(req.date, "bill date")
        validate_date(req.payment_date, "payment date")
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))

    # Use atomic counter to prevent race conditions on bill ref generation
    try:
        parts = req.date.split("-")
        date_suffix = f"{parts[2]}{parts[1]}{parts[0][2:]}"
    except Exception:
        date_suffix = "000000"

    counter_key = f"bill_seq_{req.date}"
    counter_doc = await db.counters.find_one_and_update(
        {"_id": counter_key},
        {"$inc": {"seq": 1}},
        upsert=True,
        return_document=True  # Return document after update
    )
    seq = counter_doc.get("seq", 1) if counter_doc else 1
    ref = f"{seq:02d}/{date_suffix}"
    modes_str = ", ".join(req.payment_modes) if req.payment_modes else "Cash"
    tailoring_status = "Awaiting Order" if req.needs_tailoring else "N/A"

    grand_total = 0
    for item in req.items:
        discounted_price = round(item.price - (item.price * item.discount / 100), 0)
        grand_total += round(discounted_price * item.qty, 0)

    # No hard block: amount_paid may be less than, equal to, or greater than grand_total.
    # Any amount received marks the section as Settled; pending stores the actual difference.

    items_to_insert = []
    running_paid = 0
    running_discount = 0

    for idx, item in enumerate(req.items):
        discounted_price = round(item.price - (item.price * item.discount / 100), 0)
        item_total = round(discounted_price * item.qty, 0)

        # Resolve per-item tailoring fields sent from NewBill frontend
        item_article_type   = item.article_type    or "N/A"
        item_order_no       = item.order_no        or "N/A"
        item_delivery_date  = item.delivery_date   or "N/A"
        item_emb_status     = item.embroidery_status or "N/A"

        # Resolve per-item tailoring status
        if item_order_no != "N/A":
            item_tailoring_status = "Pending"
        else:
            item_tailoring_status = tailoring_status

        # Resolve addon amount and description from line item addons
        item_addons = item.addons or []
        item_addon_amount = round(sum(float(a.get("price", 0)) for a in item_addons), 0)
        item_addon_desc = ", ".join(a.get("name", "") for a in item_addons) if item_addons else "N/A"
        item_addon_pay_mode = "Pending" if item_addon_amount > 0 else "N/A"
        item_addon_pending  = item_addon_amount if item_addon_amount > 0 else 0

        # is_settled is only meaningful when amount_paid > 0.
        # If the user ticked "Settled" but entered ₹0, treat as Pending — nothing was received,
        # so marking as Settled would hide the bill from the Settlements page permanently.
        effective_settled = req.is_settled and req.amount_paid > 0

        if effective_settled:
            total_diff = grand_total - req.amount_paid
            if idx == len(req.items) - 1:
                item_discount = total_diff - running_discount
                item_paid = req.amount_paid - running_paid
            else:
                item_discount = round(item_total * (total_diff / grand_total), 0) if grand_total > 0 else 0
                item_paid = round(item_total * (req.amount_paid / grand_total), 0) if grand_total > 0 else 0
                running_discount += item_discount
                running_paid += item_paid

            doc = {
                "id": str(uuid.uuid4()),
                "date": req.date,
                "name": req.customer_name,
                "ref": ref,
                "barcode": item.barcode,
                "price": item.price,
                "qty": item.qty,
                "discount": item.discount,
                "fabric_amount": item_total,
                "tailoring_status": item_tailoring_status,
                "article_type": item_article_type,
                "order_no": item_order_no,
                "delivery_date": item_delivery_date,
                "tailoring_amount": 0,
                "embroidery_status": item_emb_status,
                "embroidery_amount": 0,
                "addon_desc": item_addon_desc,
                "addon_amount": item_addon_amount,
                "fabric_pay_mode": f"Settled - {modes_str}",
                "fabric_pay_date": req.payment_date,
                "fabric_pending": item_discount,
                "fabric_received": item_paid,
                "labour_amount": 0,
                "labour_paid": "N/A",
                "labour_pay_date": "N/A",
                "tailoring_pay_mode": "N/A",
                "tailoring_pay_date": "N/A",
                "tailoring_received": 0,
                "tailoring_pending": 0,
                "embroidery_pay_mode": "N/A",
                "embroidery_pay_date": "N/A",
                "embroidery_received": 0,
                "embroidery_pending": 0,
                "addon_pay_mode": item_addon_pay_mode,
                "addon_pay_date": "N/A",
                "addon_received": 0,
                "addon_pending": item_addon_pending,
                "karigar": "N/A",
                "tally_fabric": False,
                "tally_tailoring": False,
                "tally_embroidery": False,
                "tally_addon": False,
                "created_at": datetime.now(timezone.utc).isoformat(),
            }
        else:
            doc = {
                "id": str(uuid.uuid4()),
                "date": req.date,
                "name": req.customer_name,
                "ref": ref,
                "barcode": item.barcode,
                "price": item.price,
                "qty": item.qty,
                "discount": item.discount,
                "fabric_amount": item_total,
                "tailoring_status": item_tailoring_status,
                "article_type": item_article_type,
                "order_no": item_order_no,
                "delivery_date": item_delivery_date,
                "tailoring_amount": 0,
                "embroidery_status": item_emb_status,
                "embroidery_amount": 0,
                "addon_desc": item_addon_desc,
                "addon_amount": item_addon_amount,
                "fabric_pay_mode": "Pending",
                "fabric_pay_date": "N/A",
                "fabric_pending": item_total,
                "fabric_received": 0,
                "labour_amount": 0,
                "labour_paid": "N/A",
                "labour_pay_date": "N/A",
                "tailoring_pay_mode": "N/A",
                "tailoring_pay_date": "N/A",
                "tailoring_received": 0,
                "tailoring_pending": 0,
                "embroidery_pay_mode": "N/A",
                "embroidery_pay_date": "N/A",
                "embroidery_received": 0,
                "embroidery_pending": 0,
                "addon_pay_mode": item_addon_pay_mode,
                "addon_pay_date": "N/A",
                "addon_received": 0,
                "addon_pending": item_addon_pending,
                "karigar": "N/A",
                "tally_fabric": False,
                "tally_tailoring": False,
                "tally_embroidery": False,
                "tally_addon": False,
                "created_at": datetime.now(timezone.utc).isoformat(),
            }

        items_to_insert.append(doc)

    if items_to_insert:
        await db.items.insert_many(items_to_insert)

    if not req.is_settled and req.amount_paid > 0:
        adv = {
            "id": str(uuid.uuid4()),
            "date": req.payment_date,
            "name": req.customer_name,
            "ref": ref,
            "amount": req.amount_paid,
            "mode": modes_str,
            "tally": False,
            "created_at": datetime.now(timezone.utc).isoformat(),
        }
        await db.advances.insert_one(adv)

    await audit_log(db, "create", current_user, "bill", ref, {"customer": req.customer_name, "items": len(items_to_insert), "total": grand_total})
    return {"message": "Bill created", "ref": ref, "items_count": len(items_to_insert), "grand_total": grand_total}

