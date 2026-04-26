#!/usr/bin/env python3
"""
Backend API Testing for VBA Retail Management System
Tests all API endpoints for the fabric/tailoring business management system
"""

import requests
import sys
import json
import os
from datetime import datetime, date
from typing import Dict, Any, List


def normalize_base_url(base_url: str) -> str:
    """Accept either host root or /api URL and normalize to host root."""
    clean = (base_url or "").strip().rstrip("/")
    if clean.endswith("/api"):
        clean = clean[:-4]
    return clean

class RetailAPITester:
    def __init__(self, base_url: str = "http://127.0.0.1:8001"):
        self.base_url = normalize_base_url(base_url)
        self.api_base = f"{self.base_url}/api"
        self.tests_run = 0
        self.tests_passed = 0
        self.failed_tests = []
        self.latest_ref = None
        self.session = requests.Session()
        self.session.headers.update({'Content-Type': 'application/json'})
        self._token: str | None = None

    def login(self, username: str = "admin", password: str = "admin123") -> bool:
        """Authenticate and store JWT so all subsequent requests are authorised."""
        # Allow main() to override default credentials via _pending_credentials
        if hasattr(self, "_pending_credentials"):
            username, password = self._pending_credentials
        try:
            url = f"{self.api_base}/auth/login"
            resp = requests.post(url, json={"username": username, "password": password},
                                 headers={"Content-Type": "application/json"})
            if resp.status_code == 200:
                token = resp.json().get("access_token")
                if token:
                    self._token = token
                    self.session.headers.update({"Authorization": f"Bearer {token}"})
                    self.log_test("Auth Login", True, f"Logged in as '{username}'")
                    return True
            self.log_test("Auth Login", False, f"HTTP {resp.status_code}: {resp.text[:120]}")
            return False
        except Exception as e:
            self.log_test("Auth Login", False, str(e))
            return False

    def log_test(self, name: str, success: bool, details: str = ""):
        """Log test result"""
        self.tests_run += 1
        if success:
            self.tests_passed += 1
            print(f"✅ {name}")
        else:
            self.failed_tests.append({"name": name, "details": details})
            print(f"❌ {name} - {details}")

    def test_api_endpoint(self, method: str, endpoint: str, expected_status: int = 200, 
                         data: Dict[Any, Any] = None, params: Dict[str, Any] = None) -> tuple:
        """Test a single API endpoint"""
        url = f"{self.api_base}/{endpoint.lstrip('/')}"
        
        try:
            if method.upper() == 'GET':
                response = self.session.get(url, params=params)
            elif method.upper() == 'POST':
                response = self.session.post(url, json=data, params=params)
            elif method.upper() == 'PUT':
                response = self.session.put(url, json=data, params=params)
            elif method.upper() == 'DELETE':
                response = self.session.delete(url, params=params)
            else:
                return False, f"Unsupported method: {method}"

            success = response.status_code == expected_status
            if success:
                try:
                    return True, response.json()
                except:
                    return True, response.text
            else:
                return False, f"Status {response.status_code}, expected {expected_status}. Response: {response.text[:200]}"
                
        except Exception as e:
            return False, f"Request failed: {str(e)}"

    def test_health_check(self):
        """Test basic API health"""
        success, result = self.test_api_endpoint('GET', '/')
        self.log_test("API Health Check", success, str(result) if not success else "")
        return success

    def test_seed_data(self):
        """Test data seeding"""
        success, result = self.test_api_endpoint('POST', '/seed')
        self.log_test("Seed Data", success, str(result) if not success else "")
        if success and isinstance(result, dict):
            print(f"   📊 Items: {result.get('items_count', 'N/A')}, Advances: {result.get('advances_count', 'N/A')}")
        return success

    def test_dashboard(self):
        """Test dashboard endpoint"""
        success, result = self.test_api_endpoint('GET', '/dashboard')
        self.log_test("Dashboard API", success, str(result) if not success else "")
        
        if success and isinstance(result, dict):
            required_fields = ['total_items', 'total_advances', 'fabric_pending_amount', 
                             'tailoring_pending_amount', 'unique_customers', 'total_revenue']
            missing_fields = [f for f in required_fields if f not in result]
            if missing_fields:
                self.log_test("Dashboard Data Structure", False, f"Missing fields: {missing_fields}")
            else:
                self.log_test("Dashboard Data Structure", True)
                print(f"   📊 Items: {result['total_items']}, Customers: {result['unique_customers']}, Revenue: ₹{result['total_revenue']}")
        
        return success

    def test_customers(self):
        """Test customers endpoint"""
        success, result = self.test_api_endpoint('GET', '/customers')
        self.log_test("Customers API", success, str(result) if not success else "")
        
        if success and isinstance(result, list):
            print(f"   👥 Found {len(result)} customers")
            return len(result) > 0
        return success

    def test_items(self):
        """Test items endpoint"""
        success, result = self.test_api_endpoint('GET', '/items')
        self.log_test("Items API", success, str(result) if not success else "")
        
        if success and isinstance(result, dict) and 'items' in result:
            items = result['items']
            total = result.get('total', 0)
            print(f"   📦 Found {len(items)} items (total: {total})")
            
            # Test with filters
            success2, result2 = self.test_api_endpoint('GET', '/items', params={'limit': 5})
            self.log_test("Items API with Limit", success2, str(result2) if not success2 else "")
            
            return len(items) > 0
        return success

    def test_tailoring_awaiting(self):
        """Test tailoring awaiting orders"""
        success, result = self.test_api_endpoint('GET', '/tailoring/awaiting')
        self.log_test("Tailoring Awaiting Orders", success, str(result) if not success else "")
        
        if success and isinstance(result, list):
            print(f"   ✂️ Found {len(result)} awaiting orders")
        return success

    def test_jobwork_endpoints(self):
        """Test job work endpoints"""
        # Test tailoring tab
        success1, result1 = self.test_api_endpoint('GET', '/jobwork', params={'tab': 'tailoring'})
        self.log_test("JobWork Tailoring Tab", success1, str(result1) if not success1 else "")
        
        if success1 and isinstance(result1, dict):
            pending = len(result1.get('pending', []))
            stitched = len(result1.get('stitched', []))
            delivered = len(result1.get('delivered', []))
            print(f"   ✂️ Tailoring - Pending: {pending}, Stitched: {stitched}, Delivered: {delivered}")
        
        # Test embroidery tab
        success2, result2 = self.test_api_endpoint('GET', '/jobwork', params={'tab': 'embroidery'})
        self.log_test("JobWork Embroidery Tab", success2, str(result2) if not success2 else "")
        
        if success2 and isinstance(result2, dict):
            required = len(result2.get('required', []))
            in_progress = len(result2.get('in_progress', []))
            finished = len(result2.get('finished', []))
            print(f"   🎨 Embroidery - Required: {required}, In Progress: {in_progress}, Finished: {finished}")
        
        # Test filters
        success3, result3 = self.test_api_endpoint('GET', '/jobwork/filters')
        self.log_test("JobWork Filters", success3, str(result3) if not success3 else "")
        
        return success1 and success2 and success3

    def test_daybook(self):
        """Test daybook endpoints"""
        success1, result1 = self.test_api_endpoint('GET', '/daybook')
        self.log_test("Daybook API", success1, str(result1) if not success1 else "")
        
        if success1 and isinstance(result1, dict):
            pending = len(result1.get('pending', []))
            reconciled = len(result1.get('reconciled', []))
            print(f"   📚 Daybook - Pending: {pending}, Reconciled: {reconciled}")
        
        # Test dates endpoint
        success2, result2 = self.test_api_endpoint('GET', '/daybook/dates')
        self.log_test("Daybook Dates", success2, str(result2) if not success2 else "")
        
        if success2 and isinstance(result2, list):
            print(f"   📅 Found {len(result2)} unique dates")
        
        return success1 and success2

    def test_labour(self):
        """Test labour endpoints"""
        success1, result1 = self.test_api_endpoint('GET', '/labour')
        self.log_test("Labour Items", success1, str(result1) if not success1 else "")
        
        if success1 and isinstance(result1, list):
            print(f"   👷 Found {len(result1)} labour items")
        
        # Test karigars
        success2, result2 = self.test_api_endpoint('GET', '/labour/karigars')
        self.log_test("Labour Karigars", success2, str(result2) if not success2 else "")
        
        if success2 and isinstance(result2, list):
            print(f"   👨‍🎨 Found {len(result2)} karigars")
        
        return success1 and success2

    def test_settlements(self):
        """Test settlements endpoints"""
        # Test balances (should work with empty params)
        success1, result1 = self.test_api_endpoint('GET', '/settlements/balances')
        self.log_test("Settlement Balances (empty)", success1, str(result1) if not success1 else "")
        
        return success1

    def test_advances(self):
        """Test advances endpoint"""
        success, result = self.test_api_endpoint('GET', '/advances')
        self.log_test("Advances API", success, str(result) if not success else "")
        
        if success and isinstance(result, list):
            print(f"   💰 Found {len(result)} advances")
        
        return success

    def test_orders(self):
        """Test orders endpoint"""
        success, result = self.test_api_endpoint('GET', '/orders')
        self.log_test("Orders API", success, str(result) if not success else "")
        
        if success and isinstance(result, list):
            print(f"   📋 Found {len(result)} orders")
        
        return success

    def test_create_bill_flow(self):
        """Test creating a new bill"""
        bill_data = {
            "customer_name": "Test Customer API",
            "date": date.today().isoformat(),
            "payment_date": date.today().isoformat(),
            "items": [
                {
                    "barcode": "TEST001",
                    "qty": 2.5,
                    "price": 1000,
                    "discount": 5
                }
            ],
            "payment_modes": ["Cash"],
            "amount_paid": 2375,
            "is_settled": True,
            "needs_tailoring": False
        }
        
        success, result = self.test_api_endpoint('POST', '/bills', data=bill_data)
        self.log_test("Create Bill", success, str(result) if not success else "")
        
        if success and isinstance(result, dict):
            self.latest_ref = result.get('ref')
            print(f"   🧾 Bill created - Ref: {result.get('ref')}, Total: ₹{result.get('grand_total')}")
        
        return success

    def test_item_edit_delete(self):
        """Test item editing and deletion functionality"""
        # First get an item to edit
        success, result = self.test_api_endpoint('GET', '/items', params={'limit': 1})
        if not success or not isinstance(result, dict) or not result.get('items'):
            self.log_test("Item Edit/Delete Setup", False, "No items found to test with")
            return False
        
        item = result['items'][0]
        item_id = item['id']
        original_price = item.get('price', 0)
        
        # Test item update
        update_data = {
            "price": original_price + 100,
            "qty": 2.0,
            "discount": 10.0
        }
        
        success1, result1 = self.test_api_endpoint('PUT', f'/items/{item_id}', data=update_data)
        self.log_test("Item Update (PUT)", success1, str(result1) if not success1 else "")
        
        if success1 and isinstance(result1, dict):
            # Verify fabric_amount was recalculated
            expected_fabric = (update_data['price'] - (update_data['price'] * update_data['discount'] / 100)) * update_data['qty']
            actual_fabric = result1.get('fabric_amount', 0)
            if abs(actual_fabric - expected_fabric) < 1:  # Allow small rounding differences
                self.log_test("Item Update Calculation", True)
                print(f"   💰 Fabric amount recalculated: ₹{actual_fabric}")
            else:
                self.log_test("Item Update Calculation", False, f"Expected ₹{expected_fabric}, got ₹{actual_fabric}")
        
        # Test item deletion (create a test item first to avoid deleting real data)
        test_bill_data = {
            "customer_name": "DELETE_TEST_CUSTOMER",
            "date": date.today().isoformat(),
            "payment_date": date.today().isoformat(),
            "items": [{"barcode": "DELETE_TEST", "qty": 1, "price": 100, "discount": 0}],
            "payment_modes": ["Cash"],
            "amount_paid": 0,
            "is_settled": False,
            "needs_tailoring": False
        }
        
        success2, bill_result = self.test_api_endpoint('POST', '/bills', data=test_bill_data)
        if success2:
            # Get the created item to delete
            success3, items_result = self.test_api_endpoint('GET', '/items', params={'name': 'DELETE_TEST_CUSTOMER'})
            if success3 and items_result.get('items'):
                test_item_id = items_result['items'][0]['id']
                success4, delete_result = self.test_api_endpoint('DELETE', f'/items/{test_item_id}')
                self.log_test("Item Delete", success4, str(delete_result) if not success4 else "")
                
                # Verify item was deleted
                success5, verify_result = self.test_api_endpoint('GET', f'/items/{test_item_id}', expected_status=404)
                self.log_test("Item Delete Verification", success5, str(verify_result) if not success5 else "")
                
                return success1 and success4 and success5
        
        return success1

    def test_pdf_invoice(self):
        """Test PDF invoice generation"""
        # First get a reference number to test with
        success, result = self.test_api_endpoint('GET', '/items', params={'limit': 1})
        if not success or not isinstance(result, dict) or not result.get('items'):
            self.log_test("PDF Invoice Setup", False, "No items found to test with")
            return False
        
        ref = result['items'][0].get('ref')
        if not ref:
            self.log_test("PDF Invoice Setup", False, "No reference found")
            return False
        
        # Test PDF generation
        url = f"{self.api_base}/invoice?ref={ref}"
        try:
            response = self.session.get(url)
            success = response.status_code == 200 and response.headers.get('content-type') == 'application/pdf'
            
            if success:
                pdf_size = len(response.content)
                self.log_test("PDF Invoice Generation", True)
                print(f"   📄 PDF generated for ref {ref}, size: {pdf_size} bytes")
            else:
                self.log_test("PDF Invoice Generation", False, 
                            f"Status: {response.status_code}, Content-Type: {response.headers.get('content-type')}")
            
            return success
            
        except Exception as e:
            self.log_test("PDF Invoice Generation", False, str(e))
            return False

    def test_search_functionality(self):
        """Test search functionality"""
        # Test basic search
        success1, result1 = self.test_api_endpoint('GET', '/search', params={'q': 'Ambala', 'limit': 10})
        self.log_test("Search by Query", success1, str(result1) if not success1 else "")
        
        if success1 and isinstance(result1, dict):
            items = result1.get('items', [])
            total = result1.get('total', 0)
            print(f"   🔍 Search 'Ambala' found {len(items)} items (total: {total})")
        
        # Test search with filters
        success2, result2 = self.test_api_endpoint('GET', '/search', params={
            'q': '',
            'customer': 'Cash',
            'payment_status': 'Settled',
            'limit': 10
        })
        self.log_test("Search with Filters", success2, str(result2) if not success2 else "")
        
        if success2 and isinstance(result2, dict):
            items = result2.get('items', [])
            total = result2.get('total', 0)
            print(f"   🔍 Filtered search found {len(items)} items (total: {total})")
        
        # Test search with date range
        success3, result3 = self.test_api_endpoint('GET', '/search', params={
            'q': '',
            'date_from': '2024-01-01',
            'date_to': '2024-12-31',
            'limit': 5
        })
        self.log_test("Search with Date Range", success3, str(result3) if not success3 else "")
        
        return success1 and success2 and success3

    def test_reports_functionality(self):
        """Test reports and analytics functionality"""
        # Test summary report
        success1, result1 = self.test_api_endpoint('GET', '/reports/summary')
        self.log_test("Reports Summary", success1, str(result1) if not success1 else "")
        
        if success1 and isinstance(result1, dict):
            required_fields = ['total_fabric', 'total_items', 'payment_modes', 'article_types']
            missing_fields = [f for f in required_fields if f not in result1]
            if missing_fields:
                self.log_test("Reports Summary Structure", False, f"Missing fields: {missing_fields}")
            else:
                self.log_test("Reports Summary Structure", True)
                print(f"   📊 Summary - Items: {result1['total_items']}, Fabric: ₹{result1['total_fabric']}")
        
        # Test daily revenue report
        success2, result2 = self.test_api_endpoint('GET', '/reports/revenue', params={'period': 'daily'})
        self.log_test("Revenue Report Daily", success2, str(result2) if not success2 else "")
        
        if success2 and isinstance(result2, list):
            print(f"   📈 Daily revenue data: {len(result2)} entries")
        
        # Test monthly revenue report
        success3, result3 = self.test_api_endpoint('GET', '/reports/revenue', params={'period': 'monthly'})
        self.log_test("Revenue Report Monthly", success3, str(result3) if not success3 else "")
        
        if success3 and isinstance(result3, list):
            print(f"   📈 Monthly revenue data: {len(result3)} entries")
        
        # Test customer report
        success4, result4 = self.test_api_endpoint('GET', '/reports/customers')
        self.log_test("Customer Report", success4, str(result4) if not success4 else "")
        
        if success4 and isinstance(result4, list):
            print(f"   👥 Customer ranking: {len(result4)} customers")
            if result4:
                top_customer = result4[0]
                print(f"   🏆 Top customer: {top_customer.get('name')} - ₹{top_customer.get('total_fabric', 0)}")
        
        return success1 and success2 and success3 and success4

    def test_db_stats(self):
        """Test database stats endpoint"""
        success, result = self.test_api_endpoint('GET', '/db/stats')
        self.log_test("Database Stats API", success, str(result) if not success else "")
        
        if success and isinstance(result, dict):
            required_fields = ['items_count', 'advances_count']
            missing_fields = [f for f in required_fields if f not in result]
            if missing_fields:
                self.log_test("DB Stats Structure", False, f"Missing fields: {missing_fields}")
            else:
                self.log_test("DB Stats Structure", True)
                print(f"   📊 DB Stats - Items: {result['items_count']}, Advances: {result['advances_count']}")
        
        return success

    def test_excel_export(self):
        """Test Excel export functionality"""
        url = f"{self.api_base}/export/excel"
        try:
            response = self.session.get(url)
            success = response.status_code == 200
            
            if success:
                # Check if it's actually an Excel file
                content_type = response.headers.get('content-type', '')
                content_disposition = response.headers.get('content-disposition', '')
                
                is_excel = (
                    'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet' in content_type or
                    '.xlsx' in content_disposition
                )
                
                if is_excel:
                    file_size = len(response.content)
                    self.log_test("Excel Export", True)
                    print(f"   📊 Excel file exported, size: {file_size} bytes")
                else:
                    self.log_test("Excel Export", False, f"Invalid content type: {content_type}")
                    success = False
            else:
                self.log_test("Excel Export", False, f"Status: {response.status_code}")
            
            return success
            
        except Exception as e:
            self.log_test("Excel Export", False, str(e))
            return False

    def test_backup_restore(self):
        """Test backup and restore functionality"""
        # Test backup creation
        backup_url = f"{self.api_base}/backup"
        try:
            backup_response = self.session.get(backup_url)
            backup_success = backup_response.status_code == 200
            
            if backup_success:
                # Check if it's a valid JSON backup
                content_type = backup_response.headers.get('content-type', '')
                is_json = 'application/json' in content_type
                
                if is_json:
                    try:
                        backup_data = backup_response.json()
                        required_fields = ['items', 'advances', 'items_count', 'advances_count']
                        missing_fields = [f for f in required_fields if f not in backup_data]
                        
                        if missing_fields:
                            self.log_test("Backup Creation", False, f"Missing fields: {missing_fields}")
                            return False
                        else:
                            backup_size = len(backup_response.content)
                            self.log_test("Backup Creation", True)
                            print(f"   💾 Backup created - Items: {backup_data['items_count']}, Advances: {backup_data['advances_count']}, Size: {backup_size} bytes")
                            
                            # Test restore functionality
                            return self._test_restore_with_backup(backup_data)
                    except Exception as e:
                        self.log_test("Backup Creation", False, f"Invalid JSON: {str(e)}")
                        return False
                else:
                    self.log_test("Backup Creation", False, f"Invalid content type: {content_type}")
                    return False
            else:
                self.log_test("Backup Creation", False, f"Status: {backup_response.status_code}")
                return False
                
        except Exception as e:
            self.log_test("Backup Creation", False, str(e))
            return False

    def _test_restore_with_backup(self, backup_data):
        """Test restore functionality with backup data"""
        import tempfile
        import json
        
        try:
            # Create a temporary backup file
            with tempfile.NamedTemporaryFile(mode='w', suffix='.json', delete=False) as f:
                json.dump(backup_data, f, indent=2, default=str)
                temp_file_path = f.name
            
            # Test restore
            restore_url = f"{self.api_base}/restore"
            with open(temp_file_path, 'rb') as f:
                files = {'file': ('test_backup.json', f, 'application/json')}
                # Remove Content-Type header for file upload
                headers = {k: v for k, v in self.session.headers.items() if k.lower() != 'content-type'}
                response = requests.post(restore_url, files=files, headers=headers)
            
            # Clean up temp file
            import os
            os.unlink(temp_file_path)
            
            success = response.status_code == 200
            if success:
                try:
                    result = response.json()
                    if 'message' in result and 'items_count' in result and 'advances_count' in result:
                        self.log_test("Backup Restore", True)
                        print(f"   🔄 Restore completed - Items: {result['items_count']}, Advances: {result['advances_count']}")
                    else:
                        self.log_test("Backup Restore", False, "Invalid response structure")
                        success = False
                except:
                    self.log_test("Backup Restore", False, "Invalid JSON response")
                    success = False
            else:
                self.log_test("Backup Restore", False, f"Status: {response.status_code}, Response: {response.text[:200]}")
            
            return success
            
        except Exception as e:
            self.log_test("Backup Restore", False, str(e))
            return False

    def test_excel_import(self):
        """Test Excel import functionality"""
        # Create a minimal test Excel file
        try:
            import openpyxl
            import tempfile
            
            # Create a test workbook
            wb = openpyxl.Workbook()
            
            # Create Item Details sheet
            ws1 = wb.active
            ws1.title = "Item Details"
            
            # Add headers (matching the expected format)
            headers = [
                "Date", "Name", "Ref.", "Items", "Price", "Qty", "Discount", "Fabric Amount",
                "Tailoring?", "Article Type", "Order No.", "Delivery Date", "Tailoring Amount",
                "Embroidery?", "Embroidery Amount", "Add-on", "Add-on Amount",
                "Fabric Payment Mode", "Fabric Payment Date", "Fabric Pending Balance", "Fabric Payment Received",
                "Labour Amount", "Labour Paid?", "Labour Payment Date",
                "Tailoring Payment Mode", "Tailoring Payment Date", "Tailoring Payment Received", "Tailoring Pending Balance",
                "Embroidery Payment Mode", "Embroidery Payment Date", "Embroidery Payment Received", "Embroidery Pending Balance",
                "Add-On Payment Mode", "Add-On Payment Date", "Add-On Payment Received", "Add-On Pending Balance", "Karigar?"
            ]
            
            for col, header in enumerate(headers, 1):
                ws1.cell(row=1, column=col, value=header)
            
            # Add a test row
            from datetime import date
            test_row = [
                date.today(), "TEST_IMPORT_CUSTOMER", "01/010125", "TEST_ITEM_001", 1000, 2.5, 5, 2375,
                "N/A", "N/A", "N/A", "N/A", 0,
                "N/A", 0, "N/A", 0,
                "Pending", "N/A", 2375, 0,
                0, "N/A", "N/A",
                "N/A", "N/A", 0, 0,
                "N/A", "N/A", 0, 0,
                "N/A", "N/A", 0, 0, "N/A"
            ]
            
            for col, value in enumerate(test_row, 1):
                ws1.cell(row=2, column=col, value=value)
            
            # Create Advances sheet
            ws2 = wb.create_sheet("Advances")
            adv_headers = ["Advance Payment Date", "Name", "Ref", "Advance Payment Amount", "Advance Payment Mode"]
            for col, header in enumerate(adv_headers, 1):
                ws2.cell(row=1, column=col, value=header)
            
            # Add a test advance
            adv_row = [date.today(), "TEST_IMPORT_CUSTOMER", "01/010125", 500, "Cash"]
            for col, value in enumerate(adv_row, 1):
                ws2.cell(row=2, column=col, value=value)
            
            # Save to temporary file
            with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
                wb.save(f.name)
                temp_file_path = f.name
            
            # Test import with append mode first (safer)
            import_url = f"{self.api_base}/import/excel?mode=append"
            with open(temp_file_path, 'rb') as f:
                files = {'file': ('test_import.xlsx', f, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
                # Remove Content-Type header for file upload
                headers = {k: v for k, v in self.session.headers.items() if k.lower() != 'content-type'}
                response = requests.post(import_url, files=files, headers=headers)
            
            # Clean up temp file
            import os
            os.unlink(temp_file_path)
            
            success = response.status_code == 200
            if success:
                try:
                    result = response.json()
                    if 'message' in result and 'items_count' in result and 'advances_count' in result:
                        self.log_test("Excel Import", True)
                        print(f"   📊 Import completed - Items: {result['items_count']}, Advances: {result['advances_count']}")
                    else:
                        self.log_test("Excel Import", False, "Invalid response structure")
                        success = False
                except:
                    self.log_test("Excel Import", False, "Invalid JSON response")
                    success = False
            else:
                self.log_test("Excel Import", False, f"Status: {response.status_code}, Response: {response.text[:200]}")
            
            return success
            
        except ImportError:
            self.log_test("Excel Import", False, "openpyxl not available for testing")
            return False
        except Exception as e:
            self.log_test("Excel Import", False, str(e))
            return False

    def test_new_features_specific(self):
        """Test specific new features mentioned in the review request"""
        print("\n🎯 Testing Specific New Features from Review Request:")
        print("-" * 50)

        ref_to_test = self.latest_ref
        if not ref_to_test:
            self.log_test("Specific feature setup", False, "No reference available from create bill flow")
            return False
        
        # Test POST /api/jobwork/move-emb endpoint
        success1, result1 = self.test_api_endpoint('POST', '/jobwork/move-emb', data={
            "item_ids": ["test-id"],
            "new_status": "In Progress",
            "emb_labour_amount": 500,
            "emb_customer_amount": 1000
        })
        self.log_test("POST /api/jobwork/move-emb endpoint exists", success1, str(result1) if not success1 else "")
        
        # Test PDF invoice with created reference
        success2, result2 = self.test_api_endpoint('GET', '/invoice', params={'ref': ref_to_test})
        if success2:
            self.log_test(f"PDF Invoice for ref {ref_to_test}", True)
            print("   📄 PDF generated successfully")
        else:
            self.log_test(f"PDF Invoice for ref {ref_to_test}", False, str(result2))
        
        # Test settlement balances for the same reference
        success3, result3 = self.test_api_endpoint('GET', '/settlements/balances', params={'ref': ref_to_test})
        self.log_test(f"Settlement balances for ref {ref_to_test}", success3, str(result3) if not success3 else "")
        
        if success3 and isinstance(result3, dict):
            fabric_balance = result3.get('fabric', 0)
            tailoring_balance = result3.get('tailoring', 0)
            print(f"   💰 Settlement balances - Fabric: ₹{fabric_balance}, Tailoring: ₹{tailoring_balance}")
            
            has_fields = all(k in result3 for k in ["fabric", "tailoring", "embroidery", "addon", "advance"])
            self.log_test("Settlement balances include required fields", has_fields, "Missing one or more required balance keys")
        
        return success1 and success2 and success3

    def run_all_tests(self):
        """Run all API tests"""
        print("🚀 Starting VBA Retail Management API Tests")
        print(f"🌐 Testing against: {self.base_url}")
        print("=" * 60)

        # Authenticate first — all routes require a valid JWT
        if not self.login():
            print("❌ Cannot proceed without authentication. Check credentials.")
            return False

        # Core API tests
        self.test_health_check()
        self.test_seed_data()
        
        # Data retrieval tests
        self.test_dashboard()
        self.test_customers()
        self.test_items()
        self.test_tailoring_awaiting()
        self.test_jobwork_endpoints()
        self.test_daybook()
        self.test_labour()
        self.test_settlements()
        self.test_advances()
        self.test_orders()
        
        # Data creation test
        self.test_create_bill_flow()
        
        # NEW FEATURES TESTING
        print("\n🆕 Testing New Features:")
        print("-" * 40)
        self.test_item_edit_delete()
        self.test_pdf_invoice()
        self.test_search_functionality()
        self.test_reports_functionality()
        
        # DATA MANAGEMENT FEATURES
        print("\n📊 Testing Data Management Features:")
        print("-" * 40)
        self.test_db_stats()
        self.test_excel_export()
        self.test_backup_restore()
        self.test_excel_import()
        
        # SPECIFIC NEW FEATURES FROM REVIEW REQUEST
        self.test_new_features_specific()
        
        # Print summary
        print("\n" + "=" * 60)
        print(f"📊 Test Summary: {self.tests_passed}/{self.tests_run} passed")
        
        if self.failed_tests:
            print("\n❌ Failed Tests:")
            for test in self.failed_tests:
                print(f"   • {test['name']}: {test['details']}")
        
        success_rate = (self.tests_passed / self.tests_run * 100) if self.tests_run > 0 else 0
        print(f"✨ Success Rate: {success_rate:.1f}%")
        
        return self.tests_passed == self.tests_run

def main():
    """Main test runner"""
    base_url  = sys.argv[1] if len(sys.argv) > 1 else os.environ.get("RETAIL_API_BASE_URL", "http://127.0.0.1:8001")
    username  = sys.argv[2] if len(sys.argv) > 2 else os.environ.get("RETAIL_TEST_USER", "admin")
    password  = sys.argv[3] if len(sys.argv) > 3 else os.environ.get("RETAIL_TEST_PASS", "admin123")
    tester = RetailAPITester(base_url)
    tester._pending_credentials = (username, password)

    try:
        success = tester.run_all_tests()
        return 0 if success else 1
    except KeyboardInterrupt:
        print("\n⚠️ Tests interrupted by user")
        return 1
    except Exception as e:
        print(f"\n💥 Test runner failed: {e}")
        return 1

if __name__ == "__main__":
    sys.exit(main())